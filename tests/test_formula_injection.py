"""Regression tests: spreadsheet formula injection prevention.

Verifies that cell values beginning with formula-trigger characters
(=, +, -, @) are written as literal text (prefixed with a single-quote)
and NOT interpreted as active Excel formulas.
"""
from __future__ import annotations

import io

import openpyxl
import pytest

from services.export_utils import sanitize_cell


# ---------------------------------------------------------------------------
# Unit tests for sanitize_cell
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("value,expected", [
    # Formula triggers must be escaped
    ("=1+1",             "'=1+1"),
    ("=SUM(A1:A9)",      "'=SUM(A1:A9)"),
    ("+SUM(A1:A9)",      "'+SUM(A1:A9)"),
    ("-1",               "'-1"),
    ("@SUM",             "'@SUM"),
    # Normal strings are unchanged
    ("web-server-01",    "web-server-01"),
    ("Windows Server",   "Windows Server"),
    ("",                 ""),
    ("normal value",     "normal value"),
    # Non-string values pass through untouched
    (None,               None),
    (1024,               1024),
    (3.14,               3.14),
    (True,               True),
    (False,              False),
])
def test_sanitize_cell_unit(value, expected):
    assert sanitize_cell(value) == expected


# ---------------------------------------------------------------------------
# Integration test: formula value round-trips as literal text in an xlsx file
# ---------------------------------------------------------------------------

def _make_workbook_with_value(value: object) -> bytes:
    """Write a single cell workbook with the given value and return bytes."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append([sanitize_cell(value)])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_formula_injection_writes_as_literal_text():
    """=1+1 must not be stored as a formula — it must appear as the text '=1+1."""
    raw_value = "=1+1"
    xlsx_bytes = _make_workbook_with_value(raw_value)

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active
    cell_value = ws.cell(row=1, column=1).value

    # The cell must NOT contain the raw formula string triggering execution
    assert cell_value != raw_value, (
        f"Formula injection found: cell value is '{cell_value}' — expected literal text"
    )
    # The prefixed literal form is what we expect
    assert cell_value == f"'{raw_value}", (
        f"Expected \"'=1+1\" but got {cell_value!r}"
    )


def test_plus_formula_injection_writes_as_literal():
    raw_value = "+SUM(A1:A9)"
    xlsx_bytes = _make_workbook_with_value(raw_value)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    cell_value = wb.active.cell(row=1, column=1).value
    assert cell_value != raw_value
    assert cell_value == f"'{raw_value}"


def test_normal_vm_name_is_unchanged():
    """Normal VM names that don't start with formula chars must not be altered."""
    raw_value = "prod-web-server-01"
    xlsx_bytes = _make_workbook_with_value(raw_value)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    cell_value = wb.active.cell(row=1, column=1).value
    assert cell_value == raw_value


def test_numeric_value_is_unchanged():
    """Numeric values must pass through sanitize_cell unchanged."""
    raw_value = 1024
    xlsx_bytes = _make_workbook_with_value(raw_value)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    cell_value = wb.active.cell(row=1, column=1).value
    assert cell_value == raw_value
