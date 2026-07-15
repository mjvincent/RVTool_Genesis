"""Unit tests for fill_powervs_price_estimator() in pricing_template_filler.

Builds a minimal in-memory template workbook (with the required header rows)
and verifies that the filler writes correct values to the correct columns.

Column mapping under test:
  Col B (2)  — LPAR name
  Col C (3)  — Qty = 1
  Col D (4)  — Data Center
  Col E (5)  — System (machine type)
  Col F (6)  — Processor Type (S / D)
  Col G (7)  — Desired Cores
  Col H (8)  — Memory (GB)
  Col N (14) — OS (AIX / Linux / IBM i)
  Col P (16) — Storage Tier 1 (GB)

Data rows begin at row 19.
"""
import sys
import os
import io

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "api"))

import openpyxl
import pytest
from services.pricing_template_filler import fill_powervs_price_estimator, _normalize_os


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_template() -> bytes:
    """Build a minimal in-memory template with the required sheet name and
    placeholder header rows 17-18 (same structure as the real estimator)."""
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"   # not the target sheet — forces lookup by name
    ws = wb.create_sheet("Multiple LPAR Price Estimate")

    # Row 17 — primary header row (col B label etc.)
    ws["B17"] = "LPAR"
    ws["D17"] = "Data Center"
    ws["E17"] = "System"
    ws["G17"] = "Desired Cores"
    ws["H17"] = "Memory (GB)"

    # Row 18 — sub-header row
    ws["B18"] = "name or #"
    ws["C18"] = "LPAR Qty"

    # Row 19 — example/placeholder row (should be cleared by the filler)
    ws["B19"] = "EXAMPLE"
    ws["C19"] = 1
    ws["D19"] = "DAL10"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _read_filled(filled_bytes: bytes, sheet_name: str = "Multiple LPAR Price Estimate"):
    """Open a filled workbook and return its active worksheet."""
    wb = openpyxl.load_workbook(io.BytesIO(filled_bytes), data_only=True)
    return wb[sheet_name]


# ---------------------------------------------------------------------------
# OS normalization
# ---------------------------------------------------------------------------

class TestNormalizeOs:
    def test_aix(self):
        assert _normalize_os("AIX") == "AIX"
        assert _normalize_os("aix 7.2") == "AIX"

    def test_ibm_i(self):
        assert _normalize_os("IBM i") == "IBM i"
        assert _normalize_os("IBMi") == "IBM i"
        assert _normalize_os("OS/400") == "IBM i"
        assert _normalize_os("i5OS") == "IBM i"

    def test_vios_maps_to_aix(self):
        assert _normalize_os("VIOS") == "AIX"
        assert _normalize_os("Virtual I/O Server") == "AIX"

    def test_linux_variants(self):
        assert _normalize_os("RHEL") == "Linux"
        assert _normalize_os("Red Hat Enterprise Linux") == "Linux"
        assert _normalize_os("SLES 15") == "Linux"
        assert _normalize_os("Ubuntu 22.04") == "Linux"
        assert _normalize_os("CentOS 7") == "Linux"
        assert _normalize_os("Rocky Linux") == "Linux"
        assert _normalize_os("Linux BYOL") == "Linux"

    def test_none_defaults_to_aix(self):
        assert _normalize_os(None) == "AIX"
        assert _normalize_os("") == "AIX"

    def test_unknown_defaults_to_aix(self):
        assert _normalize_os("Windows") == "AIX"


# ---------------------------------------------------------------------------
# Column mapping
# ---------------------------------------------------------------------------

class TestColumnMapping:
    def setup_method(self):
        self.template = _make_template()

    def test_single_aix_server_correct_columns(self):
        servers = [{
            "server_name":   "msprap410",
            "machine_type":  "S1022",
            "processor_type": "S",
            "cores":         8.0,
            "memory_gb":     32.0,
            "os_config":     "AIX",
            "storage_gb":    1098,
        }]
        filled = fill_powervs_price_estimator(self.template, servers, "DAL10")
        ws = _read_filled(filled)

        assert ws.cell(19, 2).value  == "msprap410"   # B — name
        assert ws.cell(19, 3).value  == 1              # C — qty
        assert ws.cell(19, 4).value  == "DAL10"        # D — datacenter
        assert ws.cell(19, 5).value  == "S1022"        # E — system
        assert ws.cell(19, 6).value  == "S"            # F — proc type
        assert ws.cell(19, 7).value  == 8.0            # G — cores
        assert ws.cell(19, 8).value  == 32.0           # H — memory GB
        assert ws.cell(19, 14).value == "AIX"          # N — OS
        assert ws.cell(19, 16).value == 1098           # P — storage GB

    def test_linux_server_os_normalized(self):
        servers = [{"server_name": "linuxsrv", "machine_type": "S1022",
                    "cores": 4.0, "memory_gb": 16.0, "os_config": "RHEL 8",
                    "storage_gb": 200}]
        filled = fill_powervs_price_estimator(self.template, servers, "DAL10")
        ws = _read_filled(filled)
        assert ws.cell(19, 14).value == "Linux"

    def test_ibm_i_server_os_normalized(self):
        servers = [{"server_name": "ibmisrv", "machine_type": "S1022",
                    "cores": 2.0, "memory_gb": 8.0, "os_config": "IBM i",
                    "storage_gb": 500}]
        filled = fill_powervs_price_estimator(self.template, servers, "DAL10")
        ws = _read_filled(filled)
        assert ws.cell(19, 14).value == "IBM i"

    def test_vios_maps_to_aix(self):
        servers = [{"server_name": "viossrv", "machine_type": "S1022",
                    "cores": 2.0, "memory_gb": 4.0, "os_config": "VIOS",
                    "storage_gb": 100}]
        filled = fill_powervs_price_estimator(self.template, servers, "DAL10")
        ws = _read_filled(filled)
        assert ws.cell(19, 14).value == "AIX"

    def test_qty_always_one(self):
        servers = [{"server_name": "s1"}, {"server_name": "s2"}, {"server_name": "s3"}]
        filled = fill_powervs_price_estimator(self.template, servers, "DAL10")
        ws = _read_filled(filled)
        for row_num in [19, 20, 21]:
            assert ws.cell(row_num, 3).value == 1, f"Row {row_num} qty should be 1"


# ---------------------------------------------------------------------------
# Multi-server batch
# ---------------------------------------------------------------------------

class TestMultiServer:
    def test_three_servers_produce_three_rows(self):
        template = _make_template()
        servers = [
            {"server_name": "srv-a", "machine_type": "S1022", "cores": 4.0, "memory_gb": 32.0, "os_config": "AIX", "storage_gb": 100},
            {"server_name": "srv-b", "machine_type": "E1050", "cores": 8.0, "memory_gb": 64.0, "os_config": "Linux BYOL", "storage_gb": 200},
            {"server_name": "srv-c", "machine_type": "E1080", "cores": 16.0, "memory_gb": 128.0, "os_config": "IBM i", "storage_gb": 500},
        ]
        filled = fill_powervs_price_estimator(template, servers, "WDC06")
        ws = _read_filled(filled)

        assert ws.cell(19, 2).value == "srv-a"
        assert ws.cell(19, 5).value == "S1022"
        assert ws.cell(19, 14).value == "AIX"

        assert ws.cell(20, 2).value == "srv-b"
        assert ws.cell(20, 5).value == "E1050"
        assert ws.cell(20, 14).value == "Linux"

        assert ws.cell(21, 2).value == "srv-c"
        assert ws.cell(21, 5).value == "E1080"
        assert ws.cell(21, 14).value == "IBM i"

    def test_datacenter_written_to_all_rows(self):
        template = _make_template()
        servers = [{"server_name": f"s{i}"} for i in range(5)]
        filled = fill_powervs_price_estimator(template, servers, "FRA04")
        ws = _read_filled(filled)
        for row_num in range(19, 24):
            assert ws.cell(row_num, 4).value == "FRA04"


# ---------------------------------------------------------------------------
# Machine type passthrough
# ---------------------------------------------------------------------------

class TestMachineTypePassthrough:
    def test_s1022_passed_through(self):
        template = _make_template()
        filled = fill_powervs_price_estimator(template, [{"server_name": "x", "machine_type": "S1022"}], "DAL10")
        assert _read_filled(filled).cell(19, 5).value == "S1022"

    def test_e1050_passed_through(self):
        template = _make_template()
        filled = fill_powervs_price_estimator(template, [{"server_name": "x", "machine_type": "E1050"}], "DAL10")
        assert _read_filled(filled).cell(19, 5).value == "E1050"

    def test_e1080_passed_through(self):
        template = _make_template()
        filled = fill_powervs_price_estimator(template, [{"server_name": "x", "machine_type": "E1080"}], "DAL10")
        assert _read_filled(filled).cell(19, 5).value == "E1080"

    def test_s922_legacy_passed_through(self):
        template = _make_template()
        filled = fill_powervs_price_estimator(template, [{"server_name": "x", "machine_type": "S922"}], "DAL10")
        assert _read_filled(filled).cell(19, 5).value == "S922"

    def test_e980_legacy_passed_through(self):
        template = _make_template()
        filled = fill_powervs_price_estimator(template, [{"server_name": "x", "machine_type": "E980"}], "DAL10")
        assert _read_filled(filled).cell(19, 5).value == "E980"

    def test_missing_machine_type_defaults_to_s1022(self):
        template = _make_template()
        filled = fill_powervs_price_estimator(template, [{"server_name": "x"}], "DAL10")
        assert _read_filled(filled).cell(19, 5).value == "S1022"


# ---------------------------------------------------------------------------
# Memory resolution
# ---------------------------------------------------------------------------

class TestMemoryResolution:
    def test_memory_gb_field_used_directly(self):
        template = _make_template()
        filled = fill_powervs_price_estimator(template, [{"server_name": "x", "memory_gb": 64.0}], "DAL10")
        assert _read_filled(filled).cell(19, 8).value == 64.0

    def test_memory_mb_converted_to_gb(self):
        template = _make_template()
        filled = fill_powervs_price_estimator(template, [{"server_name": "x", "memory_mb": 32768}], "DAL10")
        assert _read_filled(filled).cell(19, 8).value == 32.0


# ---------------------------------------------------------------------------
# Wrong sheet name raises clear error
# ---------------------------------------------------------------------------

class TestWrongSheet:
    def test_missing_sheet_raises_value_error(self):
        wb = openpyxl.Workbook()
        wb.active.title = "WrongSheet"
        buf = io.BytesIO()
        wb.save(buf)
        template_bytes = buf.getvalue()
        with pytest.raises(ValueError, match="Multiple LPAR Price Estimate"):
            fill_powervs_price_estimator(template_bytes, [{"server_name": "x"}], "DAL10")
