"""IBM PowerVS Price Estimator template filler.

Takes a user-uploaded IBM PowerVS Price Estimator .xlsx template and writes
LPAR data from a processed PowerVS export into the correct cells of the
'Multiple LPAR Price Estimate' sheet.

Column mapping (verified against filled reference output, 2025-07-15):
  Col B (2)  — LPAR name
  Col C (3)  — LPAR Qty (always 1)
  Col D (4)  — Data Center (e.g. DAL10)
  Col E (5)  — System (machine type: S1022, E1050, E1080, S922, E980)
  Col F (6)  — Processor Type (S = Shared, D = Dedicated)
  Col G (7)  — Desired Cores (fractional e.g. 4.5 for POWER9)
  Col H (8)  — Memory (GB)
  Col N (14) — OS (AIX / Linux / IBM i)
  Col P (16) — Storage Tier 1 (GB)

Data rows begin at row 19.  The blank template contains example/placeholder
rows in rows 19-22; those are cleared before writing.
"""
from __future__ import annotations

import io
from typing import Any


def _normalize_os(os_raw: str | None) -> str:
    """Map raw OS strings to the three values the estimator accepts."""
    if not os_raw:
        return "AIX"
    os_lower = os_raw.lower().strip()
    # IBM i variants
    if any(x in os_lower for x in ("ibm i", "ibmi", "os/400", "i5os")):
        return "IBM i"
    # VIOS is AIX-based
    if "vios" in os_lower:
        return "AIX"
    # Linux variants
    if any(x in os_lower for x in (
        "linux", "rhel", "red hat", "centos", "sles", "suse",
        "rocky", "ubuntu", "debian", "fedora", "oracle linux",
    )):
        return "Linux"
    # AIX
    if "aix" in os_lower:
        return "AIX"
    # Default
    return "AIX"


def _normalize_machine_type(machine_type: str | None) -> str:
    """Pass machine type through as-is; default to S1022 if missing."""
    if not machine_type:
        return "S1022"
    return machine_type.strip()


def fill_powervs_price_estimator(
    template_bytes: bytes,
    servers: list[dict[str, Any]],
    datacenter: str = "DAL10",
) -> bytes:
    """Fill an IBM PowerVS Price Estimator template with LPAR data.

    Args:
        template_bytes: Raw bytes of the blank estimator .xlsx template.
        servers:        List of server dicts from the DB.  Each dict must have
                        at minimum the keys used in the column mapping.  Keys
                        accepted (all optional with safe defaults):
                          server_name / vm_name / name
                          machine_type
                          processor_type  ('S' or 'D'; default 'S')
                          cores / cpus / num_cpus
                          memory_gb / memory_mb
                          os_config / os_family / os
                          storage_gb / provisioned_mb / total_disk_mb
        datacenter:     Target datacenter string written to col D (e.g. 'DAL10').

    Returns:
        Raw bytes of the filled .xlsx workbook.
    """
    # Import here to keep the module importable even if openpyxl is not installed
    # (unit tests that don't test this path won't fail at import time).
    try:
        import openpyxl
    except ImportError as exc:
        raise ImportError("openpyxl is required for pricing template filling") from exc

    wb = openpyxl.load_workbook(
        io.BytesIO(template_bytes),
        keep_vba=False,
        data_only=False,   # keep formulas intact
    )

    sheet_name = "Multiple LPAR Price Estimate"
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Template does not contain the expected sheet '{sheet_name}'. "
            f"Available sheets: {wb.sheetnames}"
        )
    ws = wb[sheet_name]

    # Clear existing example/placeholder data rows (rows 19–22 in the blank template).
    # We clear up to row 30 to be safe without touching the header rows.
    _DATA_START_ROW = 19
    _CLEAR_THROUGH_ROW = 30
    for r in range(_DATA_START_ROW, _CLEAR_THROUGH_ROW + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            # Only clear cells that contain plain values (not header labels or formulas
            # in fixed rows).  We stop if the cell value looks like a header string.
            val = cell.value
            if val is not None and not _looks_like_header(val):
                cell.value = None

    # Write one row per server starting at DATA_START_ROW
    for idx, srv in enumerate(servers):
        row_num = _DATA_START_ROW + idx

        name         = _get(srv, "server_name", "vm_name", "name") or f"LPAR-{idx + 1}"
        machine_type = _normalize_machine_type(_get(srv, "machine_type"))
        proc_type    = str(_get(srv, "processor_type") or "S").strip().upper()[:1] or "S"
        cores        = _get_numeric(srv, "cores", "cpus", "num_cpus") or 0.5
        memory_gb    = _get_memory_gb(srv)
        os_raw       = _get(srv, "os_config", "os_family", "os", "os_type")
        os_val       = _normalize_os(os_raw)
        storage_gb   = _get_storage_gb(srv)

        ws.cell(row=row_num, column=2).value  = name           # B — LPAR name
        ws.cell(row=row_num, column=3).value  = 1              # C — Qty
        ws.cell(row=row_num, column=4).value  = datacenter     # D — Data Center
        ws.cell(row=row_num, column=5).value  = machine_type   # E — System
        ws.cell(row=row_num, column=6).value  = proc_type      # F — Processor Type
        ws.cell(row=row_num, column=7).value  = cores          # G — Desired Cores
        ws.cell(row=row_num, column=8).value  = memory_gb      # H — Memory (GB)
        ws.cell(row=row_num, column=14).value = os_val         # N — OS
        ws.cell(row=row_num, column=16).value = storage_gb     # P — Storage Tier 1 (GB)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _looks_like_header(val: Any) -> bool:
    """Return True if val appears to be a header/instruction string, not data."""
    if not isinstance(val, str):
        return False
    # Long instructional strings in the template's fixed rows
    return len(val) > 60 or val.startswith("*") or "Input what" in val


def _get(d: dict, *keys: str) -> Any:
    """Return the first non-None value found under any of the given keys."""
    for k in keys:
        v = d.get(k)
        if v is not None and v != "":
            return v
    return None


def _get_numeric(d: dict, *keys: str) -> float | None:
    """Return the first numeric value found under any of the given keys."""
    for k in keys:
        v = d.get(k)
        if v is not None:
            try:
                n = float(v)
                if n > 0:
                    return n
            except (TypeError, ValueError):
                pass
    return None


def _get_memory_gb(srv: dict) -> float:
    """Resolve memory in GB from any of the common field names."""
    # Try direct GB field first
    gb = _get_numeric(srv, "memory_gb")
    if gb:
        return gb
    # Fall back to memory_mb → convert
    mb = _get_numeric(srv, "memory_mb", "memory")
    if mb:
        return max(1.0, round(mb / 1024, 0))
    return 4.0  # safe default


def _get_storage_gb(srv: dict) -> int:
    """Resolve total storage in GB from any of the common field names."""
    # Try direct GB field
    gb = _get_numeric(srv, "storage_gb", "disk_gb")
    if gb:
        return int(gb)
    # Total disk in MB → GB
    mb = _get_numeric(srv, "total_disk_mb", "provisioned_mb")
    if mb:
        return max(1, round(mb / 1024))
    return 100  # safe default
