"""IBM Power Virtual Server (PowerVS) Calculator workbook generator.

Produces the 3-sheet workbook consumed by the IBM Cloud Solutioning tool for
PowerVS pricing (AIX / IBM i / Linux on Power workloads):
  - Project Settings  — Zone and per-server Compute + Storage rows
  - Exceptions        — Servers that exceed known PowerVS machine type limits
  - Data Domains      — Static reference table of valid PowerVS values

Design mirrors vpc_calculator_generator.py but uses PowerVS-specific constructs:
  - Machine types: s922 (POWER9), s1022 (POWER10), e980 (POWER9 enterprise)
  - CPU type: Shared Uncapped / Capped / Dedicated
  - Entitled processors: cpus × 0.5  (50% entitlement — documented assumption)
  - Storage: Tier 1 (AIX/IBM i, NVMe) or Tier 3 (Linux on Power, HDD)
  - No boot-disk 100/250 GB clamping — full original disk size passed through
  - OS family: PowerVS family strings (AIX, IBM i, Linux BYOL, SAP Red Hat, etc.)
"""
from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ---------------------------------------------------------------------------
# IBM PowerVS machine type selection
# Format: (max_cores, max_mem_gb, machine_type, generation)
# Rules based on IBM PowerVS server specifications:
#   s922  — POWER9, up to 15 cores, up to 960 GB RAM
#   s1022 — POWER10, up to 24 cores, up to 1920 GB RAM
#   e980  — POWER9, up to 143.5 cores, up to 15307 GB RAM (enterprise)
# ---------------------------------------------------------------------------

_PVS_MACHINE_RULES: list[tuple[int, int, str]] = [
    # (max_cores, max_mem_gb, machine_type)
    (15,  960,   "s922"),
    (24,  1920,  "s1022"),
    (143, 15307, "e980"),
]

# Entitled processors default: cpus × 0.5 (50% entitlement, documented assumption)
_ENTITLEMENT_FACTOR = 0.5
_MIN_ENTITLEMENT    = 0.5   # PowerVS minimum entitlement is 0.25; we use 0.5 as our floor


def _select_pvs_machine_type(cpus: int, mem_gb: int) -> tuple[str, str]:
    """Return (machine_type, issues_flag) for a PowerVS server.

    issues_flag is '' for a clean match or 'no_matching_profile' when the
    spec exceeds all known PowerVS machine type limits.
    """
    if cpus <= 0:
        cpus = 1
    if mem_gb <= 0:
        mem_gb = 1

    for max_cores, max_mem, machine_type in _PVS_MACHINE_RULES:
        if cpus <= max_cores and mem_gb <= max_mem:
            return machine_type, ""

    # Exceeds largest known machine type — flag it but still return e980
    return "e980", "no_matching_profile"


# ---------------------------------------------------------------------------
# Storage tier selection
# AIX / IBM i → Tier 1 (NVMe, 10 IOPS/GB — standard for mission-critical Power)
# Linux on Power → Tier 3 (HDD, 3 IOPS/GB — cost-effective)
# ---------------------------------------------------------------------------

_PVS_STORAGE_TIER: dict[str, str] = {
    "aix":         "Tier 1",
    "ibm i":       "Tier 1",
    "ibm i mol":   "Tier 1",
    "linux byol":  "Tier 3",
    "sap red hat": "Tier 3",
    "sap suse":    "Tier 3",
    "red hat gp":  "Tier 3",
    "suse gp":     "Tier 3",
}


def _map_pvs_storage_tier(os_family: str | None) -> str:
    """Return PowerVS storage tier based on OS family."""
    if os_family is None:
        return "Tier 1"
    return _PVS_STORAGE_TIER.get(os_family.lower().strip(), "Tier 1")


# ---------------------------------------------------------------------------
# IBM PowerVS regions and datacenters
# PowerVS datacenters use short names (dal10, lon06, tok04) unlike VPC zones
# ---------------------------------------------------------------------------

IBM_POWERVS_REGIONS: dict[str, str] = {
    "us-south":  "North America",
    "us-east":   "North America",
    "ca-tor":    "North America",
    "eu-de":     "Europe",
    "eu-gb":     "Europe",
    "jp-tok":    "Asia Pacific",
    "jp-osa":    "Asia Pacific",
    "au-syd":    "Asia Pacific",
    "in-che":    "Asia Pacific",
    "br-sao":    "South America",
}

IBM_POWERVS_DATACENTERS: dict[str, list[str]] = {
    "us-south":  ["dal10", "dal12"],
    "us-east":   ["wdc06", "wdc07"],
    "ca-tor":    ["tor01"],
    "eu-de":     ["fra04", "fra05"],
    "eu-gb":     ["lon04", "lon06"],
    "jp-tok":    ["tok02", "tok04"],
    "jp-osa":    ["osa21"],
    "au-syd":    ["syd04", "syd05"],
    "in-che":    ["che01"],
    "br-sao":    ["sao01", "sao04"],
}


def get_pvs_geography(region: str) -> str:
    """Return the geography name for a PowerVS region."""
    return IBM_POWERVS_REGIONS.get(region, "North America")


def get_pvs_valid_datacenters(region: str) -> list[str]:
    """Return the list of valid datacenter names for a PowerVS region."""
    return IBM_POWERVS_DATACENTERS.get(region, ["dal10"])


# ---------------------------------------------------------------------------
# Workbook styling helpers (mirrors vpc_calculator_generator)
# ---------------------------------------------------------------------------

_THIN        = Side(style="thin")
_ALL_BORDERS = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL = PatternFill("solid", fgColor="D3D3D3")
_HEADER_FONT = Font(bold=True)
_ISSUE_FILL  = PatternFill("solid", fgColor="FFF2CC")   # light yellow — flags exception rows


def _write_header(ws: Any, headers: list[str]) -> None:
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.border = _ALL_BORDERS
        cell.alignment = Alignment(wrap_text=True)


def _auto_size(ws: Any, headers: list[str]) -> None:
    for col_idx, header in enumerate(headers, 1):
        max_len = len(header)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 40)


# ---------------------------------------------------------------------------
# Project Settings column headers — PowerVS-specific
# ---------------------------------------------------------------------------

PVS_PS_HEADERS = [
    "Issues",                      # col 1  — blank or comma-separated flags
    "Server name",                 # col 2  — VM/server name
    "Machine type",                # col 3  — s922 / s1022 / e980
    "Number of instances",         # col 4  — always 1
    "CPU type",                    # col 5  — Shared Uncapped / Capped / Dedicated
    "Entitled processors",         # col 6  — cpus × 0.5 (documented assumption)
    "Memory (GB)",                 # col 7  — RAM in GB
    "OS family",                   # col 8  — AIX / IBM i / Linux BYOL / etc.
    "Storage type",                # col 9  — Tier 1 or Tier 3
    "Storage size (GB)",           # col 10 — total_disk_mb / 1024 (full original disk)
    "Requirement Type",            # col 11 — Zone / Compute / Storage
    "Geography",                   # col 12 — North America / Europe / Asia Pacific / etc.
    "Region",                      # col 13 — us-south / eu-gb / jp-tok / etc.
    "Data Center",                 # col 14 — dal10 / lon06 / tok04 / etc.
    "Network type",                # col 15 — Public / Private
    "Expected bandwidth (Gbps)",   # col 16 — default 1
]

# Column index map (1-based)
_PVS_PS = {h: i for i, h in enumerate(PVS_PS_HEADERS, 1)}


def _pvs_row(n: int) -> list[Any]:
    """Return a blank n-element list for one Project Settings row."""
    return [None] * n


def _set(row: list, header: str, val: Any) -> None:
    """Set a value by header name in a PVS PS row (mutates in place)."""
    idx = _PVS_PS[header] - 1   # 0-based
    row[idx] = val


# ---------------------------------------------------------------------------
# Data Domains — static reference table of valid PowerVS values
# Mirrors the purpose of the VPC Data Domains sheet (lookup/validation reference)
# ---------------------------------------------------------------------------

_DD_HEADERS = [
    "Region",
    "Data Center",
    "Machine type",
    "CPU type",
    "OS family",
    "Storage type",
    "Requirement Type",
    "Network type",
]

_DD_ROWS: list[list[Any]] = [
    # Region,       Data Center, Machine type, CPU type,            OS family,    Storage type, Requirement Type, Network type
    ["us-south",    "dal10",     "s922",        "Shared Uncapped",  "AIX",         "Tier 1",     "Zone",           "Public"],
    ["us-south",    "dal12",     "s1022",       "Capped",           "IBM i",       "Tier 1",     "Compute",        "Private"],
    ["us-east",     "wdc06",     "e980",        "Dedicated",        "IBM i MOL",   "Tier 3",     "Storage",        None],
    ["us-east",     "wdc07",     None,          None,               "Linux BYOL",  None,          None,             None],
    ["ca-tor",      "tor01",     None,          None,               "SAP Red Hat", None,          None,             None],
    ["eu-de",       "fra04",     None,          None,               "SAP SUSE",    None,          None,             None],
    ["eu-de",       "fra05",     None,          None,               "Red Hat GP",  None,          None,             None],
    ["eu-gb",       "lon04",     None,          None,               "SUSE GP",     None,          None,             None],
    ["eu-gb",       "lon06",     None,          None,               None,          None,          None,             None],
    ["jp-tok",      "tok02",     None,          None,               None,          None,          None,             None],
    ["jp-tok",      "tok04",     None,          None,               None,          None,          None,             None],
    ["jp-osa",      "osa21",     None,          None,               None,          None,          None,             None],
    ["au-syd",      "syd04",     None,          None,               None,          None,          None,             None],
    ["au-syd",      "syd05",     None,          None,               None,          None,          None,             None],
    ["in-che",      "che01",     None,          None,               None,          None,          None,             None],
    ["br-sao",      "sao01",     None,          None,               None,          None,          None,             None],
    ["br-sao",      "sao04",     None,          None,               None,          None,          None,             None],
]

# Pad all rows to len(_DD_HEADERS) columns
_DD_COLS = len(_DD_HEADERS)
_DD_ROWS = [row + [None] * (_DD_COLS - len(row)) for row in _DD_ROWS]


# ---------------------------------------------------------------------------
# Main generator
# ---------------------------------------------------------------------------

def generate_powervs_calculator_xlsx(
    records: list[dict],
    project_name: str,
    pvs_region: str = "us-south",
    pvs_datacenter: str = "dal10",
) -> bytes:
    """Generate a 3-sheet IBM PowerVS Calculator workbook.

    Args:
        records: List of enriched record dicts from _fetch_enriched_records().
                 Each has keys: normalized_data, server_type, is_excluded.
        project_name: Used in the workbook title / filename.
        pvs_region: IBM PowerVS target region (e.g. "us-south").
        pvs_datacenter: IBM PowerVS target datacenter (e.g. "dal10").

    Returns:
        Raw bytes of the generated .xlsx workbook.

    Notes:
        - Entitled processors = cpus × 0.5 (50% entitlement). Documented assumption.
        - Storage size = total_disk_mb (full original disk, no IBM VPC clamping applied).
        - Storage tier = Tier 1 for AIX/IBM i; Tier 3 for Linux on Power.
    """
    geography = get_pvs_geography(pvs_region)
    n = len(PVS_PS_HEADERS)

    # Filter: PowerVS non-excluded only
    active = [
        r for r in records
        if not r.get("is_excluded")
        and r.get("server_type") == "powervs"
        and r.get("normalized_data")
    ]

    wb = Workbook()

    # -------------------------------------------------------------------
    # Sheet 1: Project Settings
    # -------------------------------------------------------------------
    ps_ws = wb.active
    ps_ws.title = "Project Settings"
    _write_header(ps_ws, PVS_PS_HEADERS)

    # Zone row — identifies the target PowerVS location
    zone_row = _pvs_row(n)
    _set(zone_row, "Requirement Type",        "Zone")
    _set(zone_row, "Geography",               geography)
    _set(zone_row, "Region",                  pvs_region)
    _set(zone_row, "Data Center",             pvs_datacenter)
    _set(zone_row, "Network type",            "Private")
    _set(zone_row, "Expected bandwidth (Gbps)", 1)
    ps_ws.append(zone_row)

    exception_rows: list[list[Any]] = []

    for rec in active:
        nd    = rec["normalized_data"]
        vinfo = nd.get("vinfo") or {}

        vm_name    = vinfo.get("vm_name") or "unknown"
        cpus       = int(vinfo.get("num_cpus") or vinfo.get("cpus") or 1)
        mem_mb     = int(vinfo.get("memory_mb") or vinfo.get("memory") or 4096)
        mem_gb     = max(1, round(mem_mb / 1024))

        # Full original disk size — no IBM VPC boot clamping for PowerVS
        prov_mb  = int(vinfo.get("provisioned_mb") or 51200)   # fallback 50 GB
        total_mb = int(vinfo.get("total_disk_mb") or prov_mb)
        disk_gb  = max(1, round(total_mb / 1024))

        os_family = (
            vinfo.get("powervs_os_family")
            or vinfo.get("os_config")
            or "AIX"
        )

        machine_type, issues_flag = _select_pvs_machine_type(cpus, mem_gb)
        storage_tier = _map_pvs_storage_tier(os_family)

        # Entitled processors: cpus × 0.5, minimum 0.5
        entitled = max(_MIN_ENTITLEMENT, round(cpus * _ENTITLEMENT_FACTOR, 1))

        # Compute row
        cmp_row = _pvs_row(n)
        _set(cmp_row, "Issues",                    issues_flag)
        _set(cmp_row, "Server name",               vm_name)
        _set(cmp_row, "Machine type",              machine_type)
        _set(cmp_row, "Number of instances",       1)
        _set(cmp_row, "CPU type",                  "Shared Uncapped")
        _set(cmp_row, "Entitled processors",       entitled)
        _set(cmp_row, "Memory (GB)",               mem_gb)
        _set(cmp_row, "OS family",                 os_family)
        _set(cmp_row, "Storage type",              storage_tier)
        _set(cmp_row, "Storage size (GB)",         disk_gb)
        _set(cmp_row, "Requirement Type",          "Compute")
        _set(cmp_row, "Geography",                 geography)
        _set(cmp_row, "Region",                    pvs_region)
        _set(cmp_row, "Data Center",               pvs_datacenter)
        _set(cmp_row, "Network type",              "Private")
        ps_ws.append(cmp_row)

        if issues_flag:
            exception_rows.append(list(cmp_row))

    # Style all data rows
    for row_cells in ps_ws.iter_rows(min_row=2):
        issues_val = row_cells[0].value   # col A = Issues
        fill = _ISSUE_FILL if issues_val else None
        for cell in row_cells:
            cell.border    = _ALL_BORDERS
            cell.alignment = Alignment(wrap_text=False)
            if fill:
                cell.fill = fill

    _auto_size(ps_ws, PVS_PS_HEADERS)
    ps_ws.freeze_panes = "A2"

    # -------------------------------------------------------------------
    # Sheet 2: Exceptions (same layout as Project Settings)
    # -------------------------------------------------------------------
    ex_ws = wb.create_sheet("Exceptions")
    _write_header(ex_ws, PVS_PS_HEADERS)

    # Zone row
    ex_ws.append(list(zone_row))

    for ex_row in exception_rows:
        ex_ws.append(ex_row)

    for row_cells in ex_ws.iter_rows(min_row=2):
        for cell in row_cells:
            cell.border    = _ALL_BORDERS
            cell.alignment = Alignment(wrap_text=False)

    _auto_size(ex_ws, PVS_PS_HEADERS)
    ex_ws.freeze_panes = "A2"

    # -------------------------------------------------------------------
    # Sheet 3: Data Domains (static reference table)
    # -------------------------------------------------------------------
    dd_ws = wb.create_sheet("Data Domains")
    _write_header(dd_ws, _DD_HEADERS)

    for dd_row in _DD_ROWS:
        dd_ws.append(dd_row)
        for cell in dd_ws[dd_ws.max_row]:
            cell.border    = _ALL_BORDERS
            cell.alignment = Alignment(wrap_text=False)

    _auto_size(dd_ws, _DD_HEADERS)
    dd_ws.freeze_panes = "A2"

    # -------------------------------------------------------------------
    # Serialise to bytes
    # -------------------------------------------------------------------
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
