"""Structural validator for generated RVTools .xlsx files.

Checks that a generated file has the correct sheet names, column counts,
and header spelling before it is delivered to the IBM Cool tool.
"""
from __future__ import annotations

import io

from openpyxl import load_workbook

from services.rvtools_generator import (
    VHOST_HEADERS,
    VINFO_HEADERS,
    VNETWORK_HEADERS,
    VPARTITION_HEADERS,
)

REQUIRED_SHEETS = ["vInfo", "vNetwork", "vPartition", "vHost"]

_EXPECTED_HEADERS: dict[str, list[str]] = {
    "vInfo": VINFO_HEADERS,
    "vNetwork": VNETWORK_HEADERS,
    "vPartition": VPARTITION_HEADERS,
    "vHost": VHOST_HEADERS,
}


def validate_rvtools_xlsx(file_bytes: bytes) -> dict:
    """Validate a generated RVTools .xlsx file.

    Args:
        file_bytes: raw bytes of the .xlsx workbook.

    Returns:
        A dict with keys:
          - valid (bool): True if all checks passed.
          - sheets (list[str]): sheet names found in the workbook.
          - errors (list[str]): fatal structural violations.
          - warnings (list[str]): non-fatal issues (e.g. empty data rows).
    """
    errors: list[str] = []
    warnings: list[str] = []

    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        return {
            "valid": False,
            "sheets": [],
            "errors": [f"Could not open workbook: {exc}"],
            "warnings": [],
        }

    sheets_found = wb.sheetnames

    # ------------------------------------------------------------------
    # 1. Exactly the 4 required sheets — no more, no less.
    # ------------------------------------------------------------------
    missing = [s for s in REQUIRED_SHEETS if s not in sheets_found]
    extra = [s for s in sheets_found if s not in REQUIRED_SHEETS]

    if missing:
        errors.append(f"Missing required sheets: {missing}")
    if extra:
        errors.append(f"Unexpected extra sheets: {extra}")

    # ------------------------------------------------------------------
    # 2. Per-sheet header and data-row checks.
    # ------------------------------------------------------------------
    for sheet_name in REQUIRED_SHEETS:
        if sheet_name not in sheets_found:
            continue  # already reported as missing above

        ws = wb[sheet_name]
        expected = _EXPECTED_HEADERS[sheet_name]

        # Read the header row
        rows = list(ws.iter_rows(max_row=2, values_only=True))
        if not rows:
            errors.append(f"{sheet_name}: sheet is completely empty (no header row)")
            continue

        header_row = list(rows[0])

        # Column count
        if len(header_row) != len(expected):
            errors.append(
                f"{sheet_name}: expected {len(expected)} columns, "
                f"found {len(header_row)}"
            )
        else:
            # Header spelling (exact match including trailing spaces)
            mismatches = [
                f"col {i + 1}: expected {expected[i]!r}, got {header_row[i]!r}"
                for i, (exp, got) in enumerate(zip(expected, header_row))
                if exp != got
            ]
            if mismatches:
                errors.append(
                    f"{sheet_name}: header mismatches — " + "; ".join(mismatches)
                )

        # Data rows — warn (not error) if the sheet has no data beyond the header
        has_data = len(rows) > 1 and any(v is not None for v in rows[1])
        if not has_data:
            warnings.append(f"{sheet_name}: no data rows found (sheet is empty)")

    wb.close()

    return {
        "valid": len(errors) == 0,
        "sheets": list(sheets_found),
        "errors": errors,
        "warnings": warnings,
    }
