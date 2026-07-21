"""IBM Cloud VPC Calculator workbook generator.

Produces the 3-sheet workbook consumed by the IBM Cloud Cost Estimator
(rvtools2vpc.vmware-solutions.cloud.ibm.com / IBM VPC Calculator):
  - Project Settings  — Zone, Subnet, and per-VM Compute + Data Volume rows
  - Exceptions        — VMs with no matching IBM VPC profile (flagged no_matching_profile)
  - Data Domains      — Static reference/lookup data (174 rows, never changes per-project)

Translation logic reverse-engineered from the sample file produced by the rvtools2vpc tool
using the Windows Servers 20216 Subset dataset.
"""
from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from services.export_utils import sanitize_cell

# ---------------------------------------------------------------------------
# IBM Cloud VPC compute families (confirmed from IBM Cloud Solutioning tool, mid-2025)
#
# The solutioning tool exposes 10 families:
#   Flex-Compute, Flex-Balanced, Flex-Memory, Flex-Nano,
#   Balanced, Compute, Memory, GPU, High memory, Storage optimized
#
# Selection priority (Flex families take priority per design requirement):
#   1. Flex-Compute  (cxf, 2 GB/vCPU) — try first
#   2. Flex-Balanced (bxf, 4 GB/vCPU) — try second
#   3. Flex-Memory   (mxf, 8 GB/vCPU) — try third
#   4. no_matching_profile             — no valid CPU size exists in any family
#                                         (→ Exceptions sheet)
#
# Fixed families (Balanced/Compute/Memory/GPU/High memory/Storage optimized) are present
# in the Data Domains reference sheet but are NOT used for automatic profile selection.
#
# Profile name pattern: {prefix}-{cpu}x{ram}
# RAM must equal snap_cpu × ratio exactly.
#
# CRITICAL: Each Flex family has its OWN set of valid CPU counts.
# These are derived verbatim from the _DATA_DOMAINS_ROWS reference data below
# (IBM's published catalog). A shared CPU list CANNOT be used across families —
# bxf-12x48, bxf-20x80, bxf-24x96 are NOT real IBM VPC profiles.
#
# Valid CPU sizes per family (source: Data Domains "Compute Family VS" column):
#   cxf (Flex-Compute):  2, 4, 8, 16, 24, 32, 48, 64, 96
#   bxf (Flex-Balanced): 2, 4, 8, 16, 32, 48, 64, 96          ← no 12, 20, 24
#   mxf (Flex-Memory):   2, 4, 8, 16, 24, 32, 48, 64, 96
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Flex-Nano (nxf) — fixed profile combos (NOT a sliding ratio family).
# These are the only valid nxf profiles per IBM Cloud docs (verified 2025-07-15).
# Ordered by (vcpu, ram) so a best-fit next() scan always finds the smallest match.
# ---------------------------------------------------------------------------
_NXF_PROFILES: list[tuple[int, int, str]] = [
    (1, 1, "nxf-1x1"),
    (1, 2, "nxf-1x2"),
    (1, 4, "nxf-1x4"),
    (2, 1, "nxf-2x1"),
    (2, 2, "nxf-2x2"),
]

# Per-family CPU size lists — IBM Cloud docs, verified 2025-07-15.
# bxf DOES have 24-vCPU (bxf-24x96 is a real profile); only 12 and 20 are absent.
_CXF_CPU_SIZES = [2, 4, 8, 16, 24, 32, 48, 64]   # cxf: Flex-Compute   (max: 64)
_BXF_CPU_SIZES = [2, 4, 8, 16, 24, 32, 48, 64]   # bxf: Flex-Balanced  (max: 64; no 12, 20)
_MXF_CPU_SIZES = [2, 4, 8, 16, 24, 32, 48, 64]   # mxf: Flex-Memory    (max: 64)

# Ordered list of Flex families: (gb_per_vcpu_ratio, prefix, category_label, valid_cpu_sizes)
_FLEX_FAMILIES = [
    (2, "cxf", "Flex-Compute",  _CXF_CPU_SIZES),
    (4, "bxf", "Flex-Balanced", _BXF_CPU_SIZES),
    (8, "mxf", "Flex-Memory",   _MXF_CPU_SIZES),
]

# ---------------------------------------------------------------------------
# Fixed (non-Flex) profiles for servers that exceed Flex family maximums.
# Used when all Flex families fail (typically cpus > 64 or extreme RAM ratios).
# Source: IBM Cloud VPC x86-64 instance profiles, verified 2025-07-15.
# Sorted by vcpu asc, then ram asc — first match wins.
# ---------------------------------------------------------------------------
_FIXED_PROFILES: list[tuple[int, int, str, str]] = [
    # (vcpu, ram_gb, category, profile_name)
    # Compute
    ( 96,  192, "Compute",          "cx2-96x192"),
    # Memory
    ( 96,  768, "Memory",           "mx2-96x768"),
    # Very High Memory (ux2d — 1:28 vCPU:RAM ratio)
    (  8,  224, "Very High Memory", "ux2d-8x224"),
    ( 16,  448, "Very High Memory", "ux2d-16x448"),
    ( 36, 1008, "Very High Memory", "ux2d-36x1008"),
    ( 52, 1456, "Very High Memory", "ux2d-52x1456"),
    ( 72, 2016, "Very High Memory", "ux2d-72x2016"),
    (100, 2800, "Very High Memory", "ux2d-100x2800"),
    (112, 3072, "Very High Memory", "ux2d-112x3072"),
    # Ultra High Memory (vx2d — 1:14 vCPU:RAM ratio)
    (  4,   56, "Ultra High Memory", "vx2d-4x56"),
    (  8,  112, "Ultra High Memory", "vx2d-8x112"),
    ( 16,  224, "Ultra High Memory", "vx2d-16x224"),
    ( 28,  392, "Ultra High Memory", "vx2d-28x392"),
    ( 44,  616, "Ultra High Memory", "vx2d-44x616"),
    ( 56,  784, "Ultra High Memory", "vx2d-56x784"),
    ( 88, 1232, "Ultra High Memory", "vx2d-88x1232"),
    (144, 2016, "Ultra High Memory", "vx2d-144x2016"),
    (176, 2464, "Ultra High Memory", "vx2d-176x2464"),  # catalog maximum
]

# The largest fixed profile available — used as the assumption fallback when a
# server genuinely exceeds every catalog entry (>176 vCPU or >3072 GB RAM).
_FIXED_PROFILE_MAX = (176, 2464, "Ultra High Memory", "vx2d-176x2464")


def _select_vpc_profile(cpus: int, ram_gb: int) -> tuple[str, str, str]:
    """Return (category, profile_name, issues_flag) for the best-fit IBM VPC profile.

    Algorithm (four steps — a server is ALWAYS assigned a profile):

    STEP 1 — Flex-Nano (nxf): if cpus <= 2 AND ram_gb <= 4, find the smallest nxf
      profile where p_vcpu >= cpus AND p_ram >= ram_gb.  Returns flag="".

    STEP 2 — Flex families (cxf → bxf → mxf): for each family find the smallest
      valid CPU size >= requested, then check snap_cpu × ratio >= ram_gb.
      First family that satisfies both returns flag="".

    STEP 3 — Fixed profiles (cx2, mx2, ux2d, vx2d up to 176 vCPU): scan
      _FIXED_PROFILES for the first entry where p_vcpu >= cpus AND p_ram >= ram_gb.
      Returns flag="fixed_profile" — row appears on main sheet with an assumption note.

    STEP 4 — Assumption fallback: nothing in the catalog covers the spec.
      Return the largest available profile (vx2d-176x2464) with flag="assumption".
      Server stays on the MAIN sheet — never on the Exceptions sheet.

    The Exceptions sheet is reserved for import/parse failures only.

    Returns:
        category     — e.g. "Flex-Balanced", "Compute", "Ultra High Memory"
        profile_name — e.g. "bxf-16x64", "cx2-96x192", "vx2d-176x2464"
        issues_flag  — "" | "fixed_profile" | "assumption"
    """
    if cpus <= 0:
        cpus = 1
    if ram_gb <= 0:
        ram_gb = 1

    # STEP 1: Flex-Nano — tiny workloads below cxf minimum (<=2 vCPU, <=4 GB RAM)
    if cpus <= 2 and ram_gb <= 4:
        match = next(
            (p for p in _NXF_PROFILES if p[0] >= cpus and p[1] >= ram_gb),
            None,
        )
        if match:
            return "Flex-Nano", match[2], ""

    # STEP 2: Flex families (cxf → bxf → mxf)
    for ratio, prefix, category, cpu_sizes in _FLEX_FAMILIES:
        snap_cpu = next((c for c in cpu_sizes if c >= cpus), None)
        if snap_cpu is None:
            continue
        snap_ram = snap_cpu * ratio
        if snap_ram >= ram_gb:
            return category, f"{prefix}-{snap_cpu}x{snap_ram}", ""

    # STEP 3: Fixed profiles — covers >64 vCPU and high-memory-ratio servers
    for p_vcpu, p_ram, p_cat, p_name in _FIXED_PROFILES:
        if p_vcpu >= cpus and p_ram >= ram_gb:
            return p_cat, p_name, "fixed_profile"

    # STEP 4: Assumption fallback — spec exceeds entire catalog
    _, _, fb_cat, fb_name = _FIXED_PROFILE_MAX
    return fb_cat, fb_name, "assumption"


# ---------------------------------------------------------------------------
# OS string → IBM VPC image name mapping
# ---------------------------------------------------------------------------

_OS_IMAGE_MAP: list[tuple[str, str, str]] = [
    # (substring_match_lower, os_family_name, ibm_image_id)
    # First match wins — more specific entries must come before generic fallbacks.

    # ── SAP / SQL Server variants — must precede generic RHEL / SUSE / Windows ──
    # RHEL for SAP (PLACEHOLDER image IDs — verify against IBM Cloud catalog)
    ("red hat enterprise linux for sap",     "Red Hat Enterprise Linux for SAP", "ibm-redhat-9-2-sap-hana-amd64-3"),  # PLACEHOLDER
    ("rhel for sap",                         "Red Hat Enterprise Linux for SAP", "ibm-redhat-9-2-sap-hana-amd64-3"),  # PLACEHOLDER
    # SUSE for SAP (PLACEHOLDER image IDs — verify against IBM Cloud catalog)
    ("suse linux enterprise server for sap", "SUSE Linux Enterprise Server for SAP", "ibm-sles-15-5-sap-hana-amd64-1"),  # PLACEHOLDER
    ("sles for sap",                         "SUSE Linux Enterprise Server for SAP", "ibm-sles-15-5-sap-hana-amd64-1"),  # PLACEHOLDER
    # Windows with SQL Server (PLACEHOLDER image IDs — verify against IBM Cloud catalog)
    ("windows server with sql server",       "Windows Server with SQL Server", "ibm-windows-server-2022-sql-2022-amd64-1"),  # PLACEHOLDER
    ("microsoft windows server with sql",    "Windows Server with SQL Server", "ibm-windows-server-2022-sql-2022-amd64-1"),  # PLACEHOLDER

    # ── Windows Server ────────────────────────────────────────────────────────
    ("windows server 2022",  "Windows Server", "ibm-windows-server-2022-amd64-9"),
    ("windows server 2019",  "Windows Server", "ibm-windows-server-2019-amd64-11"),
    ("windows server 2016",  "Windows Server", "ibm-windows-server-2016-amd64-12"),
    ("windows server 2012",  "Windows Server", "ibm-windows-server-2012-r2-amd64-12"),
    ("windows server 2008",  "Windows Server", "ibm-windows-server-2008-r2-amd64-12"),
    ("windows",              "Windows Server", "ibm-windows-server-2022-amd64-9"),
    # Red Hat Enterprise Linux
    ("red hat enterprise linux 9",  "Red Hat Enterprise Linux", "ibm-redhat-9-2-minimal-amd64-2"),
    ("red hat enterprise linux 8",  "Red Hat Enterprise Linux", "ibm-redhat-8-8-minimal-amd64-2"),
    ("red hat enterprise linux 7",  "Red Hat Enterprise Linux", "ibm-redhat-7-9-minimal-amd64-11"),
    ("red hat enterprise linux",    "Red Hat Enterprise Linux", "ibm-redhat-9-2-minimal-amd64-2"),
    ("rhel 9",   "Red Hat Enterprise Linux", "ibm-redhat-9-2-minimal-amd64-2"),
    ("rhel 8",   "Red Hat Enterprise Linux", "ibm-redhat-8-8-minimal-amd64-2"),
    ("rhel 7",   "Red Hat Enterprise Linux", "ibm-redhat-7-9-minimal-amd64-11"),
    ("rhel",     "Red Hat Enterprise Linux", "ibm-redhat-9-2-minimal-amd64-2"),
    # SUSE
    ("suse linux enterprise server 15", "SUSE Linux Enterprise Server", "ibm-sles-15-5-amd64-1"),
    ("suse linux enterprise server 12", "SUSE Linux Enterprise Server", "ibm-sles-12-5-amd64-6"),
    ("suse", "SUSE Linux Enterprise Server", "ibm-sles-15-5-amd64-1"),
    ("sles", "SUSE Linux Enterprise Server", "ibm-sles-15-5-amd64-1"),
    # Ubuntu
    ("ubuntu 22", "Ubuntu Linux", "ibm-ubuntu-22-04-6-minimal-amd64-2"),
    ("ubuntu 20", "Ubuntu Linux", "ibm-ubuntu-20-04-6-minimal-amd64-4"),
    ("ubuntu 18", "Ubuntu Linux", "ibm-ubuntu-18-04-6-minimal-amd64-4"),
    ("ubuntu",    "Ubuntu Linux", "ibm-ubuntu-22-04-6-minimal-amd64-2"),
    # Debian
    ("debian 12", "Debian GNU/Linux", "ibm-debian-12-0-minimal-amd64-1"),
    ("debian 11", "Debian GNU/Linux", "ibm-debian-11-7-minimal-amd64-2"),
    ("debian",    "Debian GNU/Linux", "ibm-debian-12-0-minimal-amd64-1"),
    # CentOS / CentOS Stream
    ("centos stream 9", "CentOS Stream", "ibm-centos-stream-9-amd64-4"),
    ("centos stream 8", "CentOS Stream", "ibm-centos-stream-8-amd64-4"),
    ("centos stream",   "CentOS Stream", "ibm-centos-stream-9-amd64-4"),
    ("centos 8",  "CentOS", "ibm-centos-7-9-minimal-amd64-12"),
    ("centos 7",  "CentOS", "ibm-centos-7-9-minimal-amd64-12"),
    ("centos",    "CentOS", "ibm-centos-7-9-minimal-amd64-12"),
    # Rocky Linux
    ("rocky linux 9", "Rocky Linux", "ibm-rocky-linux-9-2-minimal-amd64-1"),
    ("rocky linux 8", "Rocky Linux", "ibm-rocky-linux-8-8-minimal-amd64-2"),
    ("rocky linux",   "Rocky Linux", "ibm-rocky-linux-9-2-minimal-amd64-1"),
    # Fedora CoreOS
    ("fedora coreos",  "Fedora CoreOS", "ibm-fedora-coreos-38-stable-2"),
    ("fedora",         "Fedora CoreOS", "ibm-fedora-coreos-38-stable-2"),
    # Oracle Linux
    ("oracle linux 9", "Red Hat Enterprise Linux", "ibm-redhat-9-2-minimal-amd64-2"),
    ("oracle linux 8", "Red Hat Enterprise Linux", "ibm-redhat-8-8-minimal-amd64-2"),
    ("oracle linux",   "Red Hat Enterprise Linux", "ibm-redhat-9-2-minimal-amd64-2"),
    # Generic Linux fallback
    ("linux",  "CentOS", "ibm-centos-7-9-minimal-amd64-12"),
]


def _map_os_to_image(os_string: str | None) -> tuple[str, str]:
    """Return (os_family_name, ibm_image_id) for a given OS string.

    Falls back to CentOS 7 (the rvtools2vpc tool default) when no match.
    """
    if not os_string:
        return "CentOS", "ibm-centos-7-9-minimal-amd64-12"
    os_lower = os_string.lower()
    for substr, family, image in _OS_IMAGE_MAP:
        if substr in os_lower:
            return family, image
    # Unrecognised — use CentOS fallback
    return "CentOS", "ibm-centos-7-9-minimal-amd64-12"


# ---------------------------------------------------------------------------
# IBM Cloud geography / region / datacenter lookup
# ---------------------------------------------------------------------------

# Maps region → geography label (used in Zone/Subnet/DataVolume rows)
IBM_VPC_REGIONS: dict[str, str] = {
    "us-south":  "North America",
    "us-east":   "North America",
    "ca-tor":    "North America",
    "ca-mon":    "North America",
    "br-sao":    "South America",
    "eu-gb":     "Europe",
    "eu-de":     "Europe",
    "eu-es":     "Europe",
    "eu-fr2":    "Europe",
    "jp-tok":    "Asia Pacific",
    "jp-osa":    "Asia Pacific",
    "au-syd":    "Asia Pacific",
    "in-che":    "Asia Pacific",
    "kr-seo":    "Asia Pacific",
    "mx-qro":    "North America",
}

# Maps region → list of availability zones (datacenters)
IBM_VPC_DATACENTERS: dict[str, list[str]] = {
    "us-south":  ["us-south-1", "us-south-2", "us-south-3"],
    "us-east":   ["us-east-1",  "us-east-2",  "us-east-3"],
    "ca-tor":    ["ca-tor-1",   "ca-tor-2",   "ca-tor-3"],
    "ca-mon":    ["ca-mon-1",   "ca-mon-2",   "ca-mon-3"],
    "br-sao":    ["br-sao-1",   "br-sao-2",   "br-sao-3"],
    "eu-gb":     ["eu-gb-1",    "eu-gb-2",    "eu-gb-3"],
    "eu-de":     ["eu-de-1",    "eu-de-2",    "eu-de-3"],
    "eu-es":     ["eu-es-1",    "eu-es-2",    "eu-es-3"],
    "eu-fr2":    ["eu-fr2-1",   "eu-fr2-2",   "eu-fr2-3"],
    "jp-tok":    ["jp-tok-1",   "jp-tok-2",   "jp-tok-3"],
    "jp-osa":    ["jp-osa-1",   "jp-osa-2",   "jp-osa-3"],
    "au-syd":    ["au-syd-1",   "au-syd-2",   "au-syd-3"],
    "in-che":    ["in-che-1"],
    "kr-seo":    ["kr-seo-1",   "kr-seo-2"],
    "mx-qro":    ["mx-qro-1",   "mx-qro-2",   "mx-qro-3"],
}


def get_geography(region: str) -> str:
    return IBM_VPC_REGIONS.get(region, "North America")


def get_valid_datacenters(region: str) -> list[str]:
    return IBM_VPC_DATACENTERS.get(region, [f"{region}-1"])


# ---------------------------------------------------------------------------
# Static Data Domains sheet (174 rows — never changes per project)
# Derived from the jonesmi@us.ibm.com reference file verbatim.
# ---------------------------------------------------------------------------

_DATA_DOMAINS_HEADERS = [
    "Region", "Data Center", "Subnet access", "Access", "Compute Server Type",
    "Compute Architecture", "Compute Category VS", "Compute Category BM",
    "Feature VS", "Feature BM", "Operating System VS", "Operating System BM",
    "Operating System Version VS", "Operating System Version BM",
    "VPN Type", "Load Balancer Type", "VPN server modes", "IOPS", "Billing Type",
    "Direct Link Type", "Direct Link Version", "Port Metering", "Routing Type",
    "Speed", "Location", "Requirement Type", "Compute Family VS", "Compute Family BM",
]

# Each row: values aligned to _DATA_DOMAINS_HEADERS (None = empty cell)
_DATA_DOMAINS_ROWS = [
    ["au-syd","au-syd-1","Public","Public","Virtual Server","x86","Flex-Balanced","Balanced","{}","Local disk","CentOS","CentOS","ibm-centos-7-9-minimal-amd64-12","ibm-centos-7-9-minimal-amd64-12","Site to Site","Application","Standalone mode",3,"PAYG","DL Connect","Direct Link 2.0","Metered","Local","50 Mbps","US","Zone","bz2e-1x4","bx2-metal-96x384"],
    ["br-sao","au-syd-2","Private","Private","Bare Metal Server","s390x","Flex-Compute","Compute","{High Bandwidth}",None,"CentOS Stream","CentOS Stream","ibm-centos-stream-9-amd64-4","ibm-centos-stream-9-amd64-4","Client to Site","Network","HA mode",5,"1 Yr Reserved","DL Dedicated","Direct Link Classic","Unmetered","Global","100 Mbps","Canada","VPN","bz2-1x4","bx2d-metal-96x384"],
    ["ca-tor","au-syd-3",None,None,None,None,"Flex-Memory","Memory","{High Bandwidth, Instance Storage}",None,"Debian GNU/Linux","Debian GNU/Linux","ibm-centos-stream-8-amd64-4","ibm-debian-11-7-minimal-amd64-2",None,None,None,10,"3 Yr Reserved","DL Exchange",None,None,None,"200 Mbps","Europe","Subnet","bz2e-2x8","cx2-metal-96x192"],
    ["eu-de","br-sao-1",None,None,None,None,"Flex-Nano","High Memory","{Instance Storage}",None,"Fedora CoreOS","Red Hat Enterprise Linux","ibm-debian-12-0-minimal-amd64-1","ibm-redhat-9-2-minimal-amd64-2",None,None,None,None,None,None,None,None,None,"500 Mbps","Asia Pacific","Load Balancer","bz2e-4x16","cx2d-metal-96x192"],
    ["eu-gb","br-sao-2",None,None,None,None,"Balanced",None,None,None,"Red Hat Enterprise Linux","Red Hat Enterprise Linux for SAP","ibm-debian-11-7-minimal-amd64-2","ibm-redhat-9-0-minimal-amd64-4",None,None,None,None,None,None,None,None,None,"1 Gbps","Brazil","Compute","bz2-4x16","mx2-metal-96x768"],
    ["eu-es","br-sao-3",None,None,None,None,"Compute",None,None,None,"Red Hat Enterprise Linux for SAP","SUSE Linux Enterprise Server","ibm-debian-10-13-minimal-amd64-4","ibm-redhat-8-8-minimal-amd64-2",None,None,None,None,None,None,None,None,None,"2 Gbps","Mexico","Data Volume","bz2e-4x16","mx2d-metal-96x768"],
    ["jp-osa","ca-tor-1",None,None,None,None,"Memory",None,None,None,"Rocky Linux","SUSE Linux Enterprise Server for SAP","ibm-fedora-coreos-38-testing-2","ibm-redhat-8-6-minimal-amd64-6",None,None,None,None,None,None,None,None,None,"5 Gbps",None,"Direct Link","bz2e-8x32","bx3-metal-48x256"],
    ["jp-tok","ca-tor-2",None,None,None,None,"GPU",None,None,None,"SUSE Linux Enterprise Server","Ubuntu Linux","ibm-fedora-coreos-38-stable-2","ibm-redhat-9-2-amd64-sap-hana-1",None,None,None,None,None,None,None,None,None,"10 Gbps",None,"File Storage","bz2-8x32","bx3d-metal-48x256"],
    ["us-east","ca-tor-3",None,None,None,None,"High memory",None,None,None,"SUSE Linux Enterprise Server for SAP","VMware vSphere","ibm-redhat-9-2-minimal-amd64-2","ibm-redhat-9-2-amd64-sap-applications-1",None,None,None,None,None,None,None,None,None,None,None,None,"bz2-16x64","bx3-metal-64x256"],
    ["us-south","eu-de-1",None,None,None,None,"Storage optimized",None,None,None,"Ubuntu Linux","Windows Server","ibm-redhat-9-0-minimal-amd64-4","ibm-redhat-9-0-amd64-sap-hana-3",None,None,None,None,None,None,None,None,None,None,None,None,"bz2e-16x64","bx3d-metal-64x256"],
    [None,"eu-de-2",None,None,None,None,None,None,None,None,"Windows Server",None,"ibm-redhat-8-8-minimal-amd64-2","ibm-redhat-9-0-amd64-sap-applications-3",None,None,None,None,None,None,None,None,None,None,None,None,"cz2e-2x4","cx3-metal-48x128"],
    [None,"eu-de-3",None,None,None,None,None,None,None,None,"Windows Server with SQL Server",None,"ibm-redhat-8-6-minimal-amd64-6","ibm-redhat-8-8-amd64-sap-hana-1",None,None,None,None,None,None,None,None,None,None,None,None,"cz2-2x4","cx3d-metal-48x128"],
    [None,"eu-gb-1",None,None,None,None,None,None,None,None,"IBM Z",None,"ibm-redhat-7-9-minimal-amd64-11","ibm-redhat-8-8-amd64-sap-applications-1",None,None,None,None,None,None,None,None,None,None,None,None,"cz2-4x8","cx3-metal-64x128"],
    [None,"eu-gb-2",None,None,None,None,None,None,None,None,"Hyper Protect",None,"ibm-redhat-9-2-amd64-sap-hana-1","ibm-redhat-8-6-amd64-sap-hana-4",None,None,None,None,None,None,None,None,None,None,None,None,"cz2e-4x8","cx3d-metal-64x128"],
    [None,"eu-gb-3",None,None,None,None,None,None,None,None,None,None,"ibm-redhat-9-2-amd64-sap-applications-1","ibm-redhat-8-6-amd64-sap-applications-4",None,None,None,None,None,None,None,None,None,None,None,None,"cz2-8x16","mx3-metal-16x128"],
    [None,"eu-es-1",None,None,None,None,None,None,None,None,None,None,"ibm-redhat-9-0-amd64-sap-hana-3","ibm-redhat-8-4-amd64-sap-hana-8",None,None,None,None,None,None,None,None,None,None,None,None,"cz2e-8x16","mx3d-metal-16x128"],
    [None,"eu-es-2",None,None,None,None,None,None,None,None,None,None,"ibm-redhat-9-0-amd64-sap-applications-3","ibm-redhat-8-4-amd64-sap-applications-8",None,None,None,None,None,None,None,None,None,None,None,None,"cz2e-16x32","mx3-metal-48x512"],
    [None,"eu-es-3",None,None,None,None,None,None,None,None,None,None,"ibm-redhat-8-8-amd64-sap-hana-1","ibm-sles-15-5-amd64-1",None,None,None,None,None,None,None,None,None,None,None,None,"cz2-16x32","mx3d-metal-48x512"],
    [None,"jp-osa-1",None,None,None,None,None,None,None,None,None,None,"ibm-redhat-8-8-amd64-sap-applications-1","ibm-sles-12-5-amd64-6",None,None,None,None,None,None,None,None,None,None,None,None,"mz2e-2x16","mx3-metal-64x512"],
    [None,"jp-osa-2",None,None,None,None,None,None,None,None,None,None,"ibm-redhat-8-6-amd64-sap-hana-4","ibm-sles-15-5-amd64-sap-hana-1",None,None,None,None,None,None,None,None,None,None,None,None,"mz2-2x16","mx3d-metal-64x512"],
    [None,"jp-osa-3",None,None,None,None,None,None,None,None,None,None,"ibm-redhat-8-6-amd64-sap-applications-4","ibm-sles-15-5-amd64-sap-applications-1",None,None,None,None,None,None,None,None,None,None,None,None,"mz2e-4x32","ox2-metal-96x1536"],
    [None,"jp-tok-1",None,None,None,None,None,None,None,None,None,None,"ibm-redhat-8-4-amd64-sap-hana-8","ibm-windows-server-2022-amd64-9",None,None,None,None,None,None,None,None,None,None,None,None,"mz2-4x32","ox2d-metal-96x1536"],
    [None,"jp-tok-2",None,None,None,None,None,None,None,None,None,None,"ibm-redhat-8-4-amd64-sap-applications-8","ibm-windows-server-2019-amd64-11",None,None,None,None,None,None,None,None,None,None,None,None,"mz2e-8x64","ux2d-metal-112x3072"],
    [None,"jp-tok-3",None,None,None,None,None,None,None,None,None,None,"ibm-sles-15-5-amd64-1","ibm-windows-server-2016-amd64-12",None,None,None,None,None,None,None,None,None,None,None,None,"mz2-8x64","vx2d-metal-96x1792"],
    [None,"us-east-1",None,None,None,None,None,None,None,None,None,None,"ibm-sles-12-5-amd64-6","ibm-windows-server-2012-r2-amd64-12",None,None,None,None,None,None,None,None,None,None,None,None,"mz2e-16x128","bx2-48x192"],
    [None,"us-east-2",None,None,None,None,None,None,None,None,None,None,"ibm-sles-15-5-amd64-sap-hana-1","ibm-windows-server-2008-r2-amd64-12",None,None,None,None,None,None,None,None,None,None,None,None,"mz2-16x128","bx2d-48x192"],
    [None,"us-east-3",None,None,None,None,None,None,None,None,None,None,"ibm-sles-15-5-amd64-sap-applications-1",None,None,None,None,None,None,None,None,None,None,None,None,None,"mz2e-32x256","cx2-8x16"],
    [None,"us-south-1",None,None,None,None,None,None,None,None,None,None,"ibm-sles-15-5-amd64-sap-applications-1",None,None,None,None,None,None,None,None,None,None,None,None,None,"mz2-32x256","cx2d-8x16"],
    [None,"us-south-2",None,None,None,None,None,None,None,None,None,None,"ibm-windows-server-2022-amd64-9",None,None,None,None,None,None,None,None,None,None,None,None,None,"ox2-2x16","cx2-16x32"],
    [None,"us-south-3",None,None,None,None,None,None,None,None,None,None,"ibm-windows-server-2019-amd64-11",None,None,None,None,None,None,None,None,None,None,None,None,None,"ox2-4x32","cx2d-16x32"],
    [None,"br-sao-1",None,None,None,None,None,None,None,None,None,None,"ibm-windows-server-2016-amd64-12",None,None,None,None,None,None,None,None,None,None,None,None,None,"ox2-8x64","cx2-32x64"],
    [None,"br-sao-2",None,None,None,None,None,None,None,None,None,None,"ibm-windows-server-2012-r2-amd64-12",None,None,None,None,None,None,None,None,None,None,None,None,None,"ox2-16x128","cx2d-32x64"],
    [None,"br-sao-3",None,None,None,None,None,None,None,None,None,None,"ibm-windows-server-2008-r2-amd64-12",None,None,None,None,None,None,None,None,None,None,None,None,None,"ox2-32x256","cx2-48x96"],
    [None,"ca-tor-1",None,None,None,None,None,None,None,None,None,None,"ibm-ubuntu-22-04-6-minimal-amd64-2",None,None,None,None,None,None,None,None,None,None,None,None,None,"ox2d-2x16","cx2d-48x96"],
    [None,"ca-tor-2",None,None,None,None,None,None,None,None,None,None,"ibm-ubuntu-20-04-6-minimal-amd64-4",None,None,None,None,None,None,None,None,None,None,None,None,None,"ox2d-4x32","cx2-64x128"],
    [None,"ca-tor-3",None,None,None,None,None,None,None,None,None,None,"ibm-ubuntu-18-04-6-minimal-amd64-4",None,None,None,None,None,None,None,None,None,None,None,None,None,"ox2d-8x64","cx2d-64x128"],
    [None,"eu-fr2-1",None,None,None,None,None,None,None,None,None,None,"ibm-rocky-linux-9-2-minimal-amd64-1",None,None,None,None,None,None,None,None,None,None,None,None,None,"ox2d-16x128","cx2-96x192"],
    [None,"eu-fr2-2",None,None,None,None,None,None,None,None,None,None,"ibm-rocky-linux-8-8-minimal-amd64-2",None,None,None,None,None,None,None,None,None,None,None,None,None,"ox2d-32x256","cx2d-96x192"],
    [None,None,None,None,None,None,None,None,None,None,None,None,"ibm-debian-12-0-minimal-amd64-1",None,None,None,None,None,None,None,None,None,None,None,None,None,"bxf-2x8","mx2-2x16"],
    [None,None,None,None,None,None,None,None,None,None,None,None,"ibm-debian-11-7-minimal-amd64-2",None,None,None,None,None,None,None,None,None,None,None,None,None,"bxf-4x16","mx2d-2x16"],
    [None,None,None,None,None,None,None,None,None,None,None,None,"ibm-debian-10-13-minimal-amd64-4",None,None,None,None,None,None,None,None,None,None,None,None,None,"bxf-8x32","mx2-4x32"],
    [None,None,None,None,None,None,None,None,None,None,None,None,"ibm-centos-7-9-minimal-amd64-12",None,None,None,None,None,None,None,None,None,None,None,None,None,"bxf-16x64","mx2d-4x32"],
    [None,None,None,None,None,None,None,None,None,None,None,None,"ibm-centos-stream-9-amd64-4",None,None,None,None,None,None,None,None,None,None,None,None,None,"bxf-32x128","mx2-8x64"],
    [None,None,None,None,None,None,None,None,None,None,None,None,"ibm-centos-stream-8-amd64-4",None,None,None,None,None,None,None,None,None,None,None,None,None,"bxf-48x192","mx2d-8x64"],
    [None,None,None,None,None,None,None,None,None,None,None,None,"ibm-fedora-coreos-38-stable-2",None,None,None,None,None,None,None,None,None,None,None,None,None,"bxf-64x256","mx2-16x128"],
    [None,None,None,None,None,None,None,None,None,None,None,None,"ibm-fedora-coreos-38-testing-2",None,None,None,None,None,None,None,None,None,None,None,None,None,"bxf-96x384","mx2d-16x128"],
    [None,None,None,None,None,None,None,None,None,None,None,None,"ibm-redhat-9-2-minimal-amd64-2",None,None,None,None,None,None,None,None,None,None,None,None,None,"cxf-2x4","mx2-32x256"],
    [None,None,None,None,None,None,None,None,None,None,None,None,"ibm-redhat-9-0-minimal-amd64-4",None,None,None,None,None,None,None,None,None,None,None,None,None,"cxf-4x8","mx2d-32x256"],
    [None,None,None,None,None,None,None,None,None,None,None,None,"ibm-redhat-8-8-minimal-amd64-2",None,None,None,None,None,None,None,None,None,None,None,None,None,"cxf-8x16","mx2-48x384"],
    [None,None,None,None,None,None,None,None,None,None,None,None,"ibm-redhat-8-6-minimal-amd64-6",None,None,None,None,None,None,None,None,None,None,None,None,None,"cxf-16x32","mx2d-48x384"],
    [None,None,None,None,None,None,None,None,None,None,None,None,"ibm-redhat-7-9-minimal-amd64-11",None,None,None,None,None,None,None,None,None,None,None,None,None,"cxf-24x48","mx2-64x512"],
    [None,None,None,None,None,None,None,None,None,None,None,None,"ibm-redhat-8-8-amd64-sap-hana-1",None,None,None,None,None,None,None,None,None,None,None,None,None,"cxf-32x64","mx2d-64x512"],
    [None,None,None,None,None,None,None,None,None,None,None,None,"ibm-redhat-8-6-amd64-sap-hana-4",None,None,None,None,None,None,None,None,None,None,None,None,None,"cxf-48x96","mx2-96x768"],
    [None,None,None,None,None,None,None,None,None,None,None,None,"ibm-redhat-8-4-amd64-sap-hana-8",None,None,None,None,None,None,None,None,None,None,None,None,None,"cxf-64x128","mx2d-96x768"],
    [None,None,None,None,None,None,None,None,None,None,None,None,"ibm-redhat-9-2-amd64-sap-hana-1",None,None,None,None,None,None,None,None,None,None,None,None,None,"cxf-96x192","gx2-8x64x1v100"],
    [None,None,None,None,None,None,None,None,None,None,None,None,"ibm-redhat-9-0-amd64-sap-hana-3",None,None,None,None,None,None,None,None,None,None,None,None,None,"mxf-2x16","gx2-16x128x1v100"],
    [None,None,None,None,None,None,None,None,None,None,None,None,"ibm-redhat-8-8-amd64-sap-applications-1",None,None,None,None,None,None,None,None,None,None,None,None,None,"mxf-4x32","gx2-32x256x2v100"],
    [None,None,None,None,None,None,None,None,None,None,None,None,"ibm-redhat-8-6-amd64-sap-applications-4",None,None,None,None,None,None,None,None,None,None,None,None,None,"mxf-8x64","gx2-80x1280x8v100"],
    [None,None,None,None,None,None,None,None,None,None,None,None,"ibm-redhat-8-4-amd64-sap-applications-8",None,None,None,None,None,None,None,None,None,None,None,None,None,"mxf-16x128","vx2d-4x56"],
    [None,None,None,None,None,None,None,None,None,None,None,None,"ibm-redhat-9-2-amd64-sap-applications-1",None,None,None,None,None,None,None,None,None,None,None,None,None,"mxf-24x192","vx2d-8x112"],
    [None,None,None,None,None,None,None,None,None,None,None,None,"ibm-redhat-9-0-amd64-sap-applications-3",None,None,None,None,None,None,None,None,None,None,None,None,None,"mxf-32x256","vx2d-16x224"],
    [None,None,None,None,None,None,None,None,None,None,None,None,"ibm-sles-15-5-amd64-1",None,None,None,None,None,None,None,None,None,None,None,None,None,"mxf-48x384","vx2d-28x392"],
    [None,None,None,None,None,None,None,None,None,None,None,None,"ibm-sles-12-5-amd64-6",None,None,None,None,None,None,None,None,None,None,None,None,None,"mxf-64x512","vx2d-44x616"],
    [None,None,None,None,None,None,None,None,None,None,None,None,"ibm-sles-15-5-amd64-sap-hana-1",None,None,None,None,None,None,None,None,None,None,None,None,None,"mxf-96x768","vx2d-56x784"],
    [None,None,None,None,None,None,None,None,None,None,None,None,"ibm-sles-15-5-amd64-sap-applications-1",None,None,None,None,None,None,None,None,None,None,None,None,None,None,"vx2d-88x1232"],
    [None,None,None,None,None,None,None,None,None,None,None,None,"ibm-ubuntu-22-04-6-minimal-amd64-2",None,None,None,None,None,None,None,None,None,None,None,None,None,None,"vx2d-144x2016"],
    [None,None,None,None,None,None,None,None,None,None,None,None,"ibm-ubuntu-20-04-6-minimal-amd64-4",None,None,None,None,None,None,None,None,None,None,None,None,None,None,"vx2d-176x2464"],
    [None,None,None,None,None,None,None,None,None,None,None,None,"ibm-ubuntu-18-04-6-minimal-amd64-4",None,None,None,None,None,None,None,None,None,None,None,None,None,None,"ux2d-8x224"],
    [None,None,None,None,None,None,None,None,None,None,None,None,"ibm-rocky-linux-9-2-minimal-amd64-1",None,None,None,None,None,None,None,None,None,None,None,None,None,None,"ux2d-16x448"],
    [None,None,None,None,None,None,None,None,None,None,None,None,"ibm-rocky-linux-8-8-minimal-amd64-2",None,None,None,None,None,None,None,None,None,None,None,None,None,None,"ux2d-36x1008"],
    [None,None,None,None,None,None,None,None,None,None,None,None,"ibm-windows-server-2022-amd64-9",None,None,None,None,None,None,None,None,None,None,None,None,None,None,"ux2d-52x1456"],
    [None,None,None,None,None,None,None,None,None,None,None,None,"ibm-windows-server-2019-amd64-11",None,None,None,None,None,None,None,None,None,None,None,None,None,None,"ux2d-72x2016"],
    [None,None,None,None,None,None,None,None,None,None,None,None,"ibm-windows-server-2016-amd64-12",None,None,None,None,None,None,None,None,None,None,None,None,None,None,"ux2d-100x2800"],
    [None,None,None,None,None,None,None,None,None,None,None,None,"ibm-windows-server-2012-r2-amd64-12",None,None,None,None,None,None,None,None,None,None,None,None,None,None,"bz2e-1x4"],
    [None,None,None,None,None,None,None,None,None,None,None,None,"ibm-windows-server-2008-r2-amd64-12",None,None,None,None,None,None,None,None,None,None,None,None,None,None,"bz2-1x4"],
    # --- Additional non-Flex VS profiles (rows 76-174) from IBM reference file ---
    # bx2 / bx2d series
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"bx2d-2x8",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"bx2-2x8",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"bx2-4x16",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"bx2d-4x16",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"bx2d-8x32",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"bx2-8x32",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"bx2d-16x64",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"bx2-16x64",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"bx2d-32x128",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"bx2-32x128",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"bx2-48x192",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"bx2d-48x192",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"bx2d-64x256",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"bx2-64x256",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"bx2-96x384",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"bx2d-96x384",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"bx2d-128x512",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"bx2-128x512",None],
    # bx3d series
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"bx3d-2x10",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"bx3d-4x20",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"bx3d-8x40",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"bx3d-16x80",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"bx3d-24x120",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"bx3d-32x160",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"bx3d-48x240",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"bx3d-64x320",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"bx3d-96x480",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"bx3d-128x640",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"bx3d-176x880",None],
    # cx2 / cx2d series
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"cx2d-2x4",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"cx2-2x4",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"cx2-4x8",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"cx2d-4x8",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"cx2-8x16",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"cx2d-8x16",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"cx2d-16x32",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"cx2-16x32",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"cx2d-32x64",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"cx2-32x64",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"cx2-48x96",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"cx2d-48x96",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"cx2-64x128",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"cx2d-64x128",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"cx2-96x192",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"cx2d-96x192",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"cx2-128x256",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"cx2d-128x256",None],
    # cx3d series
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"cx3d-2x5",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"cx3d-4x10",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"cx3d-8x20",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"cx3d-16x40",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"cx3d-24x60",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"cx3d-32x80",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"cx3d-48x120",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"cx3d-64x160",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"cx3d-96x240",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"cx3d-128x320",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"cx3d-176x440",None],
    # mx2 / mx2d series
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"mx2-2x16",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"mx2d-2x16",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"mx2d-4x32",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"mx2-4x32",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"mx2-8x64",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"mx2d-8x64",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"mx2-16x128",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"mx2d-16x128",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"mx2-32x256",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"mx2d-32x256",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"mx2-48x384",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"mx2d-48x384",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"mx2d-64x512",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"mx2-64x512",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"mx2-96x768",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"mx2d-96x768",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"mx2-128x1024",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"mx2d-128x1024",None],
    # mx3d series
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"mx3d-2x20",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"mx3d-4x40",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"mx3d-8x80",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"mx3d-16x160",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"mx3d-24x240",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"mx3d-32x320",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"mx3d-48x480",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"mx3d-64x640",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"mx3d-96x960",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"mx3d-128x1280",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"mx3d-176x1760",None],
    # ux2d series
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"ux2d-2x56",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"ux2d-4x112",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"ux2d-8x224",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"ux2d-16x448",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"ux2d-36x1008",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"ux2d-48x1344",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"ux2d-72x2016",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"ux2d-100x2800",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"ux2d-200x5600",None],
    # gx2 / gx3 series
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"gx2-8x64x1v100",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"gx2-16x128x1v100",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"gx2-16x128x2v100",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"gx2-32x256x2v100",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"gx3-16x80x1l4",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"gx3-32x160x2l4",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"gx3-64x320x4l4",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"gx3-24x120x1l40s",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"gx3-48x240x2l40s",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"gx3d-160x1792x8h100",None],
    # ox2 series
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"ox2-2x16",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"ox2-4x32",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"ox2-8x64",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"ox2-16x128",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"ox2-32x256",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"ox2-64x512",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"ox2-96x768",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"ox2-128x1024",None],
    # vx2d series
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"vx2d-2x28",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"vx2d-4x56",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"vx2d-8x112",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"vx2d-16x224",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"vx2d-44x616",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"vx2d-88x1232",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"vx2d-144x2016",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"vx2d-176x2464",None],
    # nxf Flex-Nano series (confirmed in IBM reference file Data Domains)
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"nxf-2x1",None],
    [None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,"nxf-2x2",None],
]

# Pad all rows to 28 columns (some rows have fewer entries)
_DATA_DOMAINS_ROWS = [
    row + [None] * (28 - len(row)) for row in _DATA_DOMAINS_ROWS
]


# ---------------------------------------------------------------------------
# Workbook styling helpers
# ---------------------------------------------------------------------------

_THIN = Side(style="thin")
_ALL_BORDERS = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL = PatternFill("solid", fgColor="D3D3D3")
_HEADER_FONT = Font(bold=True)
_ISSUE_FILL      = PatternFill("solid", fgColor="FFF2CC")   # light yellow — disk clamping notes
_FIXED_FILL      = PatternFill("solid", fgColor="FFF2CC")   # yellow — fixed (non-Flex) profile
_ASSUMPTION_FILL = PatternFill("solid", fgColor="FFDDC1")   # orange — spec exceeds catalog max


def _write_header(ws: Any, headers: list[str]) -> None:
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.border = _ALL_BORDERS
        cell.alignment = Alignment(wrap_text=True)


def _auto_size(ws: Any, headers: list[str]) -> None:
    for col_idx, header in enumerate(headers, 1):
        max_len = len(header)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 40)


# ---------------------------------------------------------------------------
# Project Settings column headers (47 columns — exact from sample)
# ---------------------------------------------------------------------------

PS_HEADERS = [
    "Issues", "Compute name", "Compute Category VS", "Compute Family VS",
    "Number of instances", "Billing Type", "Boot Volume Size (GB)", "IOPS",
    "Data Volume Size (GB)", "Requirement Type", "VPC name", "Geography",
    "Region", "Data Center", "Expected Internet Traffic (GB)", "VPN Type",
    "Load Balancer Type", "Expected hours per month", "VPN server modes",
    "Subnet name", "Subnet purpose", "Subnet access", "Access",
    "Expected GB per month", "Compute Architecture", "Confidential Computing",
    "Compute Server Type", "Feature BM", "Compute Category BM", "Compute Family BM",
    "Operating System BM", "Operating System Version BM", "Feature VS",
    "Operating System VS", "Operating System Version VS", "Direct Link Name",
    "Direct Link Type", "Direct Link Version", "Port Metering", "Routing Type",
    "Speed", "Location", "Transfer Charges", "High Availability",
    "File storage description", "File storage size", "File storage Max IOPS (Gbs)",
]

# Column index map (1-based)
_PS = {h: i for i, h in enumerate(PS_HEADERS, 1)}


def _ps_row(n: int) -> list[Any]:
    """Return a blank 47-element list for one Project Settings row."""
    return [None] * n


def _set(row: list, header: str, val: Any) -> None:
    """Set a value by header name in a PS row (mutates in place)."""
    idx = _PS[header] - 1   # 0-based
    row[idx] = val


# ---------------------------------------------------------------------------
# Main generator
# ---------------------------------------------------------------------------

def generate_vpc_calculator_xlsx(
    records: list[dict],
    project_name: str,
    vpc_region: str = "us-south",
    vpc_datacenter: str = "us-south-1",
) -> bytes:
    """Generate a 3-sheet IBM Cloud VPC Calculator workbook.

    Args:
        records: List of enriched record dicts from _fetch_enriched_records().
                 Each has keys: normalized_data, server_type, is_excluded.
        project_name: Used in the workbook title / filename.
        vpc_region: IBM VPC target region (e.g. "us-south").
        vpc_datacenter: IBM VPC availability zone (e.g. "us-south-1").

    Returns:
        Raw bytes of the generated .xlsx workbook.
    """
    geography = get_geography(vpc_region)

    # Filter: x86 non-excluded only (PowerVS goes to a separate export)
    active = [
        r for r in records
        if not r.get("is_excluded")
        and r.get("server_type", "vm") != "powervs"
        and r.get("normalized_data")
    ]

    wb = Workbook()

    # -------------------------------------------------------------------
    # Sheet 1: Project Settings
    # -------------------------------------------------------------------
    ps_ws = wb.active
    ps_ws.title = "Project Settings"
    _write_header(ps_ws, PS_HEADERS)

    # Zone row
    zone_row = _ps_row(len(PS_HEADERS))
    _set(zone_row, "Requirement Type", "Zone")
    _set(zone_row, "VPC name",         "Default VPC")
    _set(zone_row, "Geography",        geography)
    _set(zone_row, "Region",           vpc_region)
    _set(zone_row, "Data Center",      vpc_datacenter)
    ps_ws.append(zone_row)

    # Subnet row
    sub_row = _ps_row(len(PS_HEADERS))
    _set(sub_row, "Requirement Type", "Subnet")
    _set(sub_row, "Geography",        geography)
    _set(sub_row, "Region",           vpc_region)
    _set(sub_row, "Data Center",      vpc_datacenter)
    _set(sub_row, "Subnet name",      "Default VPC")
    _set(sub_row, "Subnet access",    "Private")
    ps_ws.append(sub_row)

    # Per-VM: Compute + Data Volume rows
    # (exception_rows reserved for future import/parse failures; not populated by profile mismatches)
    exception_rows: list[list[Any]] = []

    for rec in active:
        nd = rec["normalized_data"]
        vinfo = nd.get("vinfo") or {}
        server_type = rec.get("server_type", "vm")

        vm_name    = vinfo.get("vm_name", "unknown")
        cpus       = int(vinfo.get("num_cpus") or vinfo.get("cpus") or 1)
        mem_mb     = int(vinfo.get("memory_mb") or vinfo.get("memory") or 4096)
        mem_gb     = max(1, round(mem_mb / 1024))
        # provisioned_mb = boot disk (already IBM-VPC-clamped to 100–250 GB at normalize time).
        # total_disk_mb  = the customer's FULL original disk size before IBM VPC clamping.
        #                  This is what we use to compute the Data Volume overflow.
        #                  Falls back to provisioned_mb for records normalized before this field existed.
        prov_mb  = int(vinfo.get("provisioned_mb") or mem_mb)
        total_mb = int(vinfo.get("total_disk_mb") or prov_mb)
        prov_gb  = max(1, round(prov_mb / 1024))
        total_gb = max(1, round(total_mb / 1024))
        os_config  = vinfo.get("os_config") or vinfo.get("os_vmware_tools") or ""
        is_bm      = server_type == "bare_metal"

        os_family, os_image = _map_os_to_image(os_config)
        category, family, profile_flag = _select_vpc_profile(cpus, mem_gb)

        # IBM Cloud VPC boot volume: 100 GB minimum, 250 GB maximum.
        # provisioned_mb is already clamped at normalisation time (ai_normalizer.py),
        # but we re-apply the clamping here so the export is always correct even for
        # records normalized before this logic was deployed.
        boot_gb = max(100, min(250, prov_gb))
        # Data volume = anything beyond 250 GB, computed from the FULL original disk size.
        # Using total_gb here (not prov_gb) ensures records where provisioned_mb was
        # already clamped to 250 GB still produce the correct overflow data volume.
        data_gb = max(0, total_gb - 250)

        # Issues column: flag only when clamping actually occurred.
        # Use total_gb (full original disk) for the > 250 check so records whose
        # provisioned_mb was already clamped to 250 GB still get the flag.
        issues_parts: list[str] = []
        if total_gb < 100:
            issues_parts.append("boot_clamped_100")
        elif total_gb > 250:
            issues_parts.append("boot_clamped_250")
        # Profile flags: fixed_profile and assumption stay on main sheet with notes.
        # Neither goes to the Exceptions sheet (which is reserved for parse failures).
        if profile_flag == "fixed_profile":
            issues_parts.append("fixed_profile: non-Flex — verify ordering path")
        elif profile_flag == "assumption":
            issues_parts.append("assumption: spec exceeds catalog max — closest profile used")
        issues_str = ",".join(issues_parts)

        # Compute row
        cmp_row = _ps_row(len(PS_HEADERS))
        _set(cmp_row, "Issues",                    issues_str)
        _set(cmp_row, "Compute name",              sanitize_cell(vm_name))
        _set(cmp_row, "Number of instances",       1)
        _set(cmp_row, "Billing Type",              "PAYG")
        _set(cmp_row, "Boot Volume Size (GB)",     boot_gb)
        _set(cmp_row, "IOPS",                      3)
        _set(cmp_row, "Requirement Type",          "Compute")
        _set(cmp_row, "Data Center",               vpc_datacenter)
        _set(cmp_row, "Compute Architecture",      "x86")
        _set(cmp_row, "Confidential Computing",    "No")
        _set(cmp_row, "Compute Server Type",       "Bare Metal Server" if is_bm else "Virtual Server")
        _set(cmp_row, "Feature VS",                "{}")
        _set(cmp_row, "Operating System VS",       sanitize_cell(os_family))
        _set(cmp_row, "Operating System Version VS", sanitize_cell(os_image))
        # Always set category and family — every server now gets a profile
        _set(cmp_row, "Compute Category VS",  category)
        _set(cmp_row, "Compute Family VS",    family)

        ps_ws.append(cmp_row)
        # profile_flag rows stay on main sheet only; no Exceptions copy needed

        # Data Volume row — only written when provisioned disk exceeds the 250 GB boot cap.
        # data_gb = 0 when prov_gb <= 250, so no unnecessary data volume rows are created.
        if data_gb > 0:
            dv_row = _ps_row(len(PS_HEADERS))
            _set(dv_row, "IOPS",                  3)
            _set(dv_row, "Data Volume Size (GB)", data_gb)
            _set(dv_row, "Requirement Type",      "Data Volume")
            _set(dv_row, "Geography",             geography)
            _set(dv_row, "Region",                vpc_region)
            _set(dv_row, "Data Center",           vpc_datacenter)
            ps_ws.append(dv_row)

    # Style all data rows — three distinct fill levels
    for row_cells in ps_ws.iter_rows(min_row=2):
        issues_val = str(row_cells[0].value or "")
        if "assumption:" in issues_val:
            fill = _ASSUMPTION_FILL   # orange — spec exceeds catalog max
        elif issues_val:
            fill = _FIXED_FILL        # yellow — fixed profile or disk clamping note
        else:
            fill = None
        for cell in row_cells:
            cell.border = _ALL_BORDERS
            cell.alignment = Alignment(wrap_text=False)
            if fill:
                cell.fill = fill

    _auto_size(ps_ws, PS_HEADERS)
    ps_ws.freeze_panes = "A2"

    # -------------------------------------------------------------------
    # Sheet 2: Exceptions (same layout as Project Settings)
    # -------------------------------------------------------------------
    ex_ws = wb.create_sheet("Exceptions")
    _write_header(ex_ws, PS_HEADERS)

    # Zone + Subnet header rows (same as Project Settings)
    ex_ws.append(list(zone_row))
    ex_ws.append(list(sub_row))

    for ex_row in exception_rows:
        ex_ws.append(ex_row)

    for row_cells in ex_ws.iter_rows(min_row=2):
        for cell in row_cells:
            cell.border = _ALL_BORDERS
            cell.alignment = Alignment(wrap_text=False)

    _auto_size(ex_ws, PS_HEADERS)
    ex_ws.freeze_panes = "A2"

    # -------------------------------------------------------------------
    # Sheet 3: Data Domains (static reference table)
    # -------------------------------------------------------------------
    dd_ws = wb.create_sheet("Data Domains")
    _write_header(dd_ws, _DATA_DOMAINS_HEADERS)

    for dd_row in _DATA_DOMAINS_ROWS:
        dd_ws.append(dd_row)
        for cell in dd_ws[dd_ws.max_row]:
            cell.border = _ALL_BORDERS
            cell.alignment = Alignment(wrap_text=False)

    _auto_size(dd_ws, _DATA_DOMAINS_HEADERS)
    dd_ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
