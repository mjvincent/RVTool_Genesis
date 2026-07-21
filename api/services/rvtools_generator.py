"""RVTools-compliant .xlsx generator.

Produces a workbook with ALL standard RVTools sheets in the correct order.
Sheets with data: vInfo, vDisk, vPartition, vNetwork, vHost.
Remaining sheets are structurally present with correct headers but no data rows
(they are required by downstream tools such as VCF Migration Lite for format
validation, but carry no information we can synthesise).
"""
from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from services.export_utils import sanitize_cell


# ---------------------------------------------------------------------------
# Header definitions — exact spelling and order from RVTools 4.x
# ---------------------------------------------------------------------------

VINFO_HEADERS = [
    "VM", "Powerstate", "Template", "CPUs", "Memory", "NICs", "Disks",
    "Provisioned MB", "In Use MB", "Datacenter", "Cluster", "Host",
    "OS according to the configuration file", "OS according to the VMware Tools",
]

VCPU_HEADERS = [
    "VM", "Powerstate", "Template", "CPUs", "Cores per socket",
    "CPU Hot Add", "CPU Hot Remove", "CPU reservation (MHz)",
    "CPU limit (MHz)", "CPU shares", "Latency sensitivity",
    "Datacenter", "Cluster", "Host",
    "OS according to the configuration file", "OS according to the VMware Tools",
]

VMEMORY_HEADERS = [
    "VM", "Powerstate", "Template", "Memory", "Memory Hot Add",
    "Memory reservation (MB)", "Memory limit (MB)", "Memory shares",
    "Datacenter", "Cluster", "Host",
    "OS according to the configuration file", "OS according to the VMware Tools",
]

VDISK_HEADERS = [
    "VM", "Powerstate", "Template", "Disk", "Disk Mode",
    "Capacity MiB", "Disk Path", "Thin", "Datacenter", "Cluster", "Host",
    "OS according to the configuration file", "OS according to the VMware Tools",
]

VPARTITION_HEADERS = [
    "VM", "Powerstate", "Template", "Disk", "Capacity MB",
    "Consumed MB", "Free MB",
    "Free % ",   # trailing space matches sample exactly
    "Datacenter", "Cluster", "Host",
    "OS according to the configuration file", "OS according to the VMware Tools",
]

VNETWORK_HEADERS = [
    "VM", "Powerstate", "Template", "SRM Placeholder", "NIC label",
    "Adapter", "Network", "Switch", "Connected", "Starts Connected",
    "Mac Address", "Type", "IPv4 Address", "IPv6 Address",
    "Direct Path IO", "Internal Sort Column", "Annotation",
]

VFLOPPY_HEADERS = [
    "VM", "Powerstate", "Template", "Floppy label", "Connected",
    "Starts Connected", "Datacenter", "Cluster", "Host",
]

VCD_HEADERS = [
    "VM", "Powerstate", "Template", "CD label", "CD Path", "Connected",
    "Starts Connected", "Datacenter", "Cluster", "Host",
    "OS according to the configuration file", "OS according to the VMware Tools",
]

VSNAPSHOT_HEADERS = [
    "VM", "Powerstate", "Template", "Snapshot name", "Snapshot description",
    "Snapshot creation date", "Snapshot size (MB)", "Quiesced",
    "Datacenter", "Cluster", "Host",
]

VTOOLS_HEADERS = [
    "VM", "Powerstate", "Template", "DNS Name", "Primary IP Address",
    "Network", "HW version", "VMware Tools version", "VMware Tools status",
    "VMware Tools running status", "VMware Tools version status",
    "Datacenter", "Cluster", "Host",
    "OS according to the configuration file", "OS according to the VMware Tools",
]

VRP_HEADERS = [
    "Name", "Parent RP", "CPU Shares", "CPU Reservation (MHz)",
    "CPU Limit (MHz)", "CPU Expandable Reservation", "Memory Shares",
    "Memory Reservation (MB)", "Memory Limit (MB)", "Memory Expandable Reservation",
    "Datacenter", "Cluster",
]

VCLUSTER_HEADERS = [
    "Name", "Datacenter", "HAEnabled", "HA Admission Control",
    "HA Failover Level", "DRS Enabled", "DRS Automation Level",
    "DRS Migration Threshold", "EVC Mode", "Config status",
    "# CPU", "# Cores", "Total CPU (MHz)", "CPU usage (MHz)",
    "CPU usage %", "# Memory", "Memory usage %", "# VMs",
    "# Hosts", "# vCPU", "# vRAM",
]

VHOST_HEADERS = [
    "Host", "Datacenter", "Cluster", "Config status", "CPU Model",
    "Speed", "HT Available", "HT Active", "# CPU", "Cores per CPU",
    "# Cores", "CPU usage %", "# Memory", "Memory usage %", "Console",
    "# NICs", "# HBAs", "# VMs", "VMs per Core", "# vCPUs",
    "vCPUs per Core", "vRAM", "VM Used memory", "VM Memory Swapped",
    "VM Memory Ballooned", "ESX Version", "Vendor", "Model",
]

VHBA_HEADERS = [
    "Host", "Datacenter", "Cluster", "HBA", "Type", "Model",
    "Driver", "WWN/IQN", "Status",
]

VNIC_HEADERS = [
    "Host", "Datacenter", "Cluster", "NIC", "Type", "Driver",
    "Mac Address", "Speed (Mbps)", "Duplex", "Status",
]

VSWITCH_HEADERS = [
    "Host", "Datacenter", "Cluster", "Name", "Ports",
    "Uplinks", "VMs", "Version", "Type",
]

VPORT_HEADERS = [
    "Host", "Datacenter", "Cluster", "Switch", "Port group",
    "VLAN", "Active uplinks", "Standby uplinks",
]

VSCVM_HEADERS = [
    "VM", "Powerstate", "Template", "SC VM", "Datacenter", "Cluster", "Host",
]

VDATASTORE_HEADERS = [
    "Name", "Type", "URL", "Accessible", "Capacity MB", "Free Space MB",
    "Uncommitted MB", "Multiple Host Access", "Datacenter", "Cluster",
]

VMULTIWRITER_HEADERS = [
    "VM", "Powerstate", "Template", "Disk", "Multi Writer",
    "Datacenter", "Cluster", "Host",
]

VHEALTH_HEADERS = [
    "VM", "Powerstate", "Template", "HW version",
    "VMware Tools version", "VMware Tools status",
    "Overall status", "Config status",
    "Datacenter", "Cluster", "Host",
]

VFILEINFO_HEADERS = [
    "VM", "Powerstate", "Template", "Config file",
    "Annotation", "Datacenter", "Cluster", "Host",
]

# Ordered list of all sheets: (sheet_name, headers)
# This order matches a real RVTools 4.x export.
ALL_SHEETS: list[tuple[str, list[str]]] = [
    ("vInfo",        VINFO_HEADERS),
    ("vCPU",         VCPU_HEADERS),
    ("vMemory",      VMEMORY_HEADERS),
    ("vDisk",        VDISK_HEADERS),
    ("vPartition",   VPARTITION_HEADERS),
    ("vNetwork",     VNETWORK_HEADERS),
    ("vFloppy",      VFLOPPY_HEADERS),
    ("vCD",          VCD_HEADERS),
    ("vSnapshot",    VSNAPSHOT_HEADERS),
    ("vTools",       VTOOLS_HEADERS),
    ("vRP",          VRP_HEADERS),
    ("vCluster",     VCLUSTER_HEADERS),
    ("vHost",        VHOST_HEADERS),
    ("vHBA",         VHBA_HEADERS),
    ("vNIC",         VNIC_HEADERS),
    ("vSwitch",      VSWITCH_HEADERS),
    ("vPort",        VPORT_HEADERS),
    ("vSC+VM",       VSCVM_HEADERS),
    ("vDatastore",   VDATASTORE_HEADERS),
    ("vMultiWriter", VMULTIWRITER_HEADERS),
    ("vHealth",      VHEALTH_HEADERS),
    ("vFileInfo",    VFILEINFO_HEADERS),
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_THIN = Side(style="thin")
_ALL_BORDERS = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL = PatternFill("solid", fgColor="D3D3D3")
_HEADER_FONT = Font(bold=True)


def _write_header(ws: Any, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.border = _ALL_BORDERS
        cell.alignment = Alignment(horizontal="center")


def _auto_size_columns(ws: Any, headers: list[str]) -> None:
    """Set column widths based on content, capped at 50."""
    col_widths = [len(h) for h in headers]
    for row in ws.iter_rows(min_row=2, values_only=True):
        for i, val in enumerate(row):
            if val is not None and i < len(col_widths):
                col_widths[i] = max(col_widths[i], len(str(val)))
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = min(width + 2, 50)


def _get(d: dict, key: str, default: Any = None) -> Any:
    return d.get(key, default)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def generate_rvtools_xlsx(
    records: list[dict],
    project_name: str,
    powervs_only: bool = False,
    x86_only: bool = False,
) -> bytes:
    """Generate a standards-compliant RVTools .xlsx from normalised server records.

    Produces all 22 standard RVTools 4.x sheets.

    Args:
        records: list of dicts with keys:
            ``normalized_data`` (the vinfo/vnetwork/etc dict),
            ``server_type`` (str),
            ``is_excluded`` (bool).
            Also accepts legacy format (plain normalized_data dicts).
        powervs_only: if True, include only server_type=="powervs" records.
        x86_only: if True, include only vm/bare_metal records (exclude powervs).
        Excluded records (is_excluded=True) are always omitted.
    """
    # Normalise to enriched format — support legacy callers that pass plain dicts
    enriched: list[dict] = []
    for r in records:
        if r is None:
            continue
        if "normalized_data" in r:
            # New enriched format: {"normalized_data": {...}, "server_type": ..., "is_excluded": ...}
            if r.get("is_excluded"):
                continue
            nd = r.get("normalized_data") or {}
            st = r.get("server_type") or (nd.get("server_type") or "vm")
        else:
            # Legacy format: raw normalized_data dict
            nd = r
            st = nd.get("server_type") or "vm"

        # Apply powervs / x86 filter
        is_powervs = (st == "powervs")
        if powervs_only and not is_powervs:
            continue
        if x86_only and is_powervs:
            continue

        enriched.append(nd)

    records = enriched  # shadow with filtered list

    wb = Workbook()
    wb.remove(wb.active)   # remove the default empty sheet

    # Create all sheets with headers in the correct order
    sheets: dict[str, Any] = {}
    for sheet_name, headers in ALL_SHEETS:
        ws = wb.create_sheet(sheet_name)
        _write_header(ws, headers)
        sheets[sheet_name] = ws

    seen_hosts: set[str] = set()

    for record in records:
        if record is None:
            continue

        vinfo       = record.get("vinfo") or {}
        vnetwork_list  = record.get("vnetwork") or []
        vpartition_list = record.get("vpartition") or []
        vhost       = record.get("vhost") or {}

        if not vinfo:
            continue

        vm_name    = sanitize_cell(_get(vinfo, "vm_name"))
        powerstate = _get(vinfo, "powerstate", "poweredOn")
        template   = _get(vinfo, "template", False)
        # For PowerVS records, use the IBM Cool OS family string in "OS according to the
        # configuration file" column so IBM Cool applies the correct pricing tier.
        # vinfo["powervs_os_family"] is set by ai_normalizer._map_powervs_os_family().
        os_cfg     = sanitize_cell(_get(vinfo, "powervs_os_family") or _get(vinfo, "os_config"))
        os_tools   = sanitize_cell(_get(vinfo, "os_vmware_tools"))
        datacenter = sanitize_cell(_get(vinfo, "datacenter"))
        cluster    = sanitize_cell(_get(vinfo, "cluster"))
        host       = sanitize_cell(_get(vinfo, "host"))
        memory_mb  = _get(vinfo, "memory_mb")
        cpus       = _get(vinfo, "cpus")

        # --- vInfo ---
        sheets["vInfo"].append([
            vm_name, powerstate, template,
            cpus, memory_mb,
            _get(vinfo, "nics"), _get(vinfo, "disks"),
            _get(vinfo, "provisioned_mb"), _get(vinfo, "in_use_mb"),
            datacenter, cluster, host, os_cfg, os_tools,
        ])

        # --- vCPU ---
        sheets["vCPU"].append([
            vm_name, powerstate, template,
            cpus,
            1,          # Cores per socket — default 1
            False,      # CPU Hot Add
            False,      # CPU Hot Remove
            0,          # CPU reservation (MHz)
            -1,         # CPU limit (-1 = unlimited)
            "normal",   # CPU shares
            "normal",   # Latency sensitivity
            datacenter, cluster, host, os_cfg, os_tools,
        ])

        # --- vMemory ---
        sheets["vMemory"].append([
            vm_name, powerstate, template,
            memory_mb,
            False,      # Memory Hot Add
            0,          # Memory reservation (MB)
            -1,         # Memory limit (-1 = unlimited)
            "normal",   # Memory shares
            datacenter, cluster, host, os_cfg, os_tools,
        ])

        # --- vDisk (one row per partition/disk) ---
        for i, part in enumerate(vpartition_list):
            if not isinstance(part, dict):
                continue
            disk_label   = sanitize_cell(_get(part, "disk_label") or f"Hard disk {i+1}")
            capacity_mb  = _get(part, "capacity_mb")
            sheets["vDisk"].append([
                vm_name, powerstate, template,
                disk_label,
                "persistent",   # Disk Mode
                capacity_mb,
                f"[datastore] {vm_name}/{vm_name}.vmdk",   # Disk Path
                True,           # Thin provisioned
                datacenter, cluster, host, os_cfg, os_tools,
            ])

        # --- vPartition ---
        for part in vpartition_list:
            if not isinstance(part, dict):
                continue
            sheets["vPartition"].append([
                vm_name, powerstate, template,
                _get(part, "disk_label"),
                _get(part, "capacity_mb"),
                _get(part, "consumed_mb"),
                _get(part, "free_mb"),
                _get(part, "free_pct"),
                datacenter, cluster, host, os_cfg, os_tools,
            ])

        # --- vNetwork ---
        for nic in vnetwork_list:
            # Guard: LLM occasionally returns strings instead of dicts in the
            # vnetwork list — skip anything that isn't a mapping.
            if not isinstance(nic, dict):
                continue
            sheets["vNetwork"].append([
                vm_name, powerstate, template,
                _get(nic, "srm_placeholder"),
                sanitize_cell(_get(nic, "nic_label")),
                sanitize_cell(_get(nic, "adapter")),
                sanitize_cell(_get(nic, "network")),
                sanitize_cell(_get(nic, "switch")),
                _get(nic, "connected"),
                _get(nic, "starts_connected"),
                sanitize_cell(_get(nic, "mac_address")),
                _get(nic, "type"),
                sanitize_cell(_get(nic, "ipv4_address")),
                sanitize_cell(_get(nic, "ipv6_address")),
                _get(nic, "direct_path_io"),
                _get(nic, "internal_sort_column"),
                sanitize_cell(_get(nic, "annotation")),
            ])

        # --- vTools (basic VM tools info) ---
        ip = None
        network_name = None
        if vnetwork_list:
            first_nic = vnetwork_list[0] or {}
            ip = _get(first_nic, "ipv4_address")
            network_name = _get(first_nic, "network")
        sheets["vTools"].append([
            vm_name, powerstate, template,
            None,           # DNS Name
            ip,
            network_name,
            "vmx-19",       # HW version (default modern)
            "11365",        # VMware Tools version (current)
            "guestToolsNotInstalled",
            "guestToolsNotRunning",
            "guestToolsNotInstalled",
            datacenter, cluster, host, os_cfg, os_tools,
        ])

        # --- vHealth ---
        sheets["vHealth"].append([
            vm_name, powerstate, template,
            "vmx-19",
            "11365",
            "guestToolsNotInstalled",
            "green",    # Overall status
            "green",    # Config status
            datacenter, cluster, host,
        ])

        # --- vFileInfo ---
        sheets["vFileInfo"].append([
            vm_name, powerstate, template,
            f"[datastore] {vm_name}/{vm_name}.vmx",
            None,       # Annotation
            datacenter, cluster, host,
        ])

        # --- vHost (deduplicated by host_name) ---
        if vhost:
            host_name = _get(vhost, "host_name") or ""
            if host_name and host_name not in seen_hosts:
                seen_hosts.add(host_name)
                sheets["vHost"].append([
                    sanitize_cell(_get(vhost, "host_name")),
                    sanitize_cell(_get(vhost, "datacenter")),
                    sanitize_cell(_get(vhost, "cluster")),
                    sanitize_cell(_get(vhost, "config_status")),
                    sanitize_cell(_get(vhost, "cpu_model")),
                    _get(vhost, "speed_mhz"),
                    _get(vhost, "ht_available"),
                    _get(vhost, "ht_active"),
                    _get(vhost, "num_cpu"),
                    _get(vhost, "cores_per_cpu"),
                    _get(vhost, "num_cores"),
                    _get(vhost, "cpu_usage_pct"),
                    _get(vhost, "memory_mb"),
                    _get(vhost, "memory_usage_pct"),
                    sanitize_cell(_get(vhost, "console")),
                    _get(vhost, "num_nics"),
                    _get(vhost, "num_hbas"),
                    _get(vhost, "num_vms"),
                    _get(vhost, "vms_per_core"),
                    _get(vhost, "num_vcpus"),
                    _get(vhost, "vcpus_per_core"),
                    _get(vhost, "vram_mb"),
                    _get(vhost, "vm_used_memory"),
                    _get(vhost, "vm_memory_swapped"),
                    _get(vhost, "vm_memory_ballooned"),
                    sanitize_cell(_get(vhost, "esx_version")),
                    sanitize_cell(_get(vhost, "vendor")),
                    sanitize_cell(_get(vhost, "model")),
                ])

    # Auto-size columns on all populated sheets
    for sheet_name, headers in ALL_SHEETS:
        _auto_size_columns(sheets[sheet_name], headers)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_rvtools_pure_xlsx(records: list[dict], project_name: str) -> bytes:
    """Generate a standard 4-sheet RVTools .xlsx from normalised server records.

    Produces exactly the four sheets a native RVTools 4.x export contains:
    vInfo, vNetwork, vPartition, vHost.

    Use this when the target tool expects a plain RVTools file rather than the
    extended 22-sheet IBM Cloud Solutioning Tool (IBM Cool / VCF Migration Lite)
    format produced by generate_rvtools_xlsx().
    """
    wb = Workbook()
    wb.remove(wb.active)

    ws_vinfo      = wb.create_sheet("vInfo")
    ws_vnetwork   = wb.create_sheet("vNetwork")
    ws_vpartition = wb.create_sheet("vPartition")
    ws_vhost      = wb.create_sheet("vHost")

    _write_header(ws_vinfo,      VINFO_HEADERS)
    _write_header(ws_vnetwork,   VNETWORK_HEADERS)
    _write_header(ws_vpartition, VPARTITION_HEADERS)
    _write_header(ws_vhost,      VHOST_HEADERS)

    seen_hosts: set[str] = set()

    for record in records:
        if record is None:
            continue

        vinfo            = record.get("vinfo") or {}
        vnetwork_list    = record.get("vnetwork") or []
        vpartition_list  = record.get("vpartition") or []
        vhost            = record.get("vhost") or {}

        # --- vInfo (one row per VM) ---
        if vinfo:
            # For PowerVS records, write the IBM Cool OS family string to the
            # "OS according to the configuration file" column.  This ensures IBM Cool
            # reads the correct pricing tier (AIX, IBM i, SAP Red Hat, etc.).
            pvs_os = _get(vinfo, "powervs_os_family") or _get(vinfo, "os_config")
            ws_vinfo.append([
                _get(vinfo, "vm_name"),
                _get(vinfo, "powerstate"),
                _get(vinfo, "template"),
                _get(vinfo, "cpus"),
                _get(vinfo, "memory_mb"),
                _get(vinfo, "nics"),
                _get(vinfo, "disks"),
                _get(vinfo, "provisioned_mb"),
                _get(vinfo, "in_use_mb"),
                _get(vinfo, "datacenter"),
                _get(vinfo, "cluster"),
                _get(vinfo, "host"),
                pvs_os,
                _get(vinfo, "os_vmware_tools"),
            ])

        # --- vNetwork (one row per NIC) ---
        for nic in vnetwork_list:
            if not isinstance(nic, dict):
                continue
            ws_vnetwork.append([
                _get(nic, "vm_name"),
                _get(nic, "powerstate"),
                _get(nic, "template"),
                _get(nic, "srm_placeholder"),
                _get(nic, "nic_label"),
                _get(nic, "adapter"),
                _get(nic, "network"),
                _get(nic, "switch"),
                _get(nic, "connected"),
                _get(nic, "starts_connected"),
                _get(nic, "mac_address"),
                _get(nic, "type"),
                _get(nic, "ipv4_address"),
                _get(nic, "ipv6_address"),
                _get(nic, "direct_path_io"),
                _get(nic, "internal_sort_column"),
                _get(nic, "annotation"),
            ])

        # --- vPartition (one row per disk/partition) ---
        for part in vpartition_list:
            if not isinstance(part, dict):
                continue
            ws_vpartition.append([
                _get(part, "vm_name"),
                _get(part, "powerstate"),
                _get(part, "template"),
                _get(part, "disk_label"),
                _get(part, "capacity_mb"),
                _get(part, "consumed_mb"),
                _get(part, "free_mb"),
                _get(part, "free_pct"),
                _get(part, "datacenter"),
                _get(part, "cluster"),
                _get(part, "host"),
                _get(part, "os_config"),
                _get(part, "os_vmware_tools"),
            ])

        # --- vHost (deduplicated by host_name) ---
        if vhost:
            host_name = _get(vhost, "host_name") or ""
            if host_name and host_name not in seen_hosts:
                seen_hosts.add(host_name)
                ws_vhost.append([
                    _get(vhost, "host_name"),
                    _get(vhost, "datacenter"),
                    _get(vhost, "cluster"),
                    _get(vhost, "config_status"),
                    _get(vhost, "cpu_model"),
                    _get(vhost, "speed_mhz"),
                    _get(vhost, "ht_available"),
                    _get(vhost, "ht_active"),
                    _get(vhost, "num_cpu"),
                    _get(vhost, "cores_per_cpu"),
                    _get(vhost, "num_cores"),
                    _get(vhost, "cpu_usage_pct"),
                    _get(vhost, "memory_mb"),
                    _get(vhost, "memory_usage_pct"),
                    _get(vhost, "console"),
                    _get(vhost, "num_nics"),
                    _get(vhost, "num_hbas"),
                    _get(vhost, "num_vms"),
                    _get(vhost, "vms_per_core"),
                    _get(vhost, "num_vcpus"),
                    _get(vhost, "vcpus_per_core"),
                    _get(vhost, "vram_mb"),
                    _get(vhost, "vm_used_memory"),
                    _get(vhost, "vm_memory_swapped"),
                    _get(vhost, "vm_memory_ballooned"),
                    _get(vhost, "esx_version"),
                    _get(vhost, "vendor"),
                    _get(vhost, "model"),
                ])

    _auto_size_columns(ws_vinfo,      VINFO_HEADERS)
    _auto_size_columns(ws_vnetwork,   VNETWORK_HEADERS)
    _auto_size_columns(ws_vpartition, VPARTITION_HEADERS)
    _auto_size_columns(ws_vhost,      VHOST_HEADERS)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
