"""Assumptions Report .xlsx generator.

Produces a two-sheet workbook:
  Sheet 1: "Assumptions" — one row per assumption, colour-coded by confidence
  Sheet 2: "Summary"     — aggregate metrics
"""
from __future__ import annotations

import io
from collections import Counter
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------

_IBM_BLUE = "0F62FE"
_WHITE = "FFFFFF"
_GREEN_FILL = PatternFill("solid", fgColor="CCFFCC")   # light green  — High
_YELLOW_FILL = PatternFill("solid", fgColor="FFFACC")  # light yellow — Medium
_RED_FILL = PatternFill("solid", fgColor="FFCCCC")     # light red    — Low

_HEADER_FILL = PatternFill("solid", fgColor=_IBM_BLUE)
_HEADER_FONT = Font(bold=True, color=_WHITE)
_BOLD_FONT = Font(bold=True)

ASSUMPTIONS_HEADERS = [
    "VM / Server Name",
    "Field Name",
    "Assumed Value",
    "Original Customer Value",
    "Reasoning",
    "Confidence",
    "Timestamp",
]

SUMMARY_HEADERS = ["Metric", "Value"]

_CONFIDENCE_FILLS = {
    "high": _GREEN_FILL,
    "medium": _YELLOW_FILL,
    "low": _RED_FILL,
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _write_header(ws: Any, headers: list[str], fill: PatternFill, font: Font) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def _auto_size_columns(ws: Any, headers: list[str]) -> None:
    col_widths = [len(h) for h in headers]
    for row in ws.iter_rows(min_row=2, values_only=True):
        for i, val in enumerate(row):
            if val is not None:
                col_widths[i] = max(col_widths[i], len(str(val)))
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = min(width + 2, 60)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

EXCLUDED_SERVERS_HEADERS = [
    "VM / Server Name",
    "Operating System",
    "Server Type",
    "Exclusion Reason",
    "Excluded At",
]

_ORANGE_FILL = PatternFill("solid", fgColor="FFE5CC")   # light orange — excluded


def generate_assumptions_xlsx(
    assumptions: list[dict],
    project_name: str,
    excluded_servers: list[dict] | None = None,
    powervs_only: bool = False,
) -> bytes:
    """Generate a human-readable Assumptions Report .xlsx.

    Args:
        assumptions: list of dicts with keys:
            vm_name, field_name, assumed_value, original_value,
            reasoning, confidence, created_at
        project_name: used for context (caller uses for filename).
        excluded_servers: optional list of excluded server records to add as a
            dedicated "Excluded Servers" sheet.  Each dict has:
            vm_name, os_config, server_type, exclusion_reason, excluded_at.
        powervs_only: if True, only include assumptions for PowerVS records.

    Returns:
        bytes: the complete Excel file ready to store or stream.
    """
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    ws_a = wb.create_sheet("Assumptions")
    ws_s = wb.create_sheet("Summary")

    # ------------------------------------------------------------------
    # Sheet 1 — Assumptions
    # ------------------------------------------------------------------
    _write_header(ws_a, ASSUMPTIONS_HEADERS, _HEADER_FILL, _HEADER_FONT)

    # Sort by VM name then field name
    sorted_assumptions = sorted(
        assumptions,
        key=lambda a: (
            str(a.get("vm_name") or "").lower(),
            str(a.get("field_name") or "").lower(),
        ),
    )

    for row_idx, a in enumerate(sorted_assumptions, start=2):
        confidence_raw = str(a.get("confidence") or "").lower()
        timestamp = a.get("created_at")
        if timestamp is not None and hasattr(timestamp, "isoformat"):
            timestamp = timestamp.isoformat()

        ws_a.append([
            a.get("vm_name"),
            a.get("field_name"),
            a.get("assumed_value"),
            a.get("original_value"),
            a.get("reasoning"),
            str(a.get("confidence") or "").capitalize(),
            timestamp,
        ])

        # Apply confidence colour to the entire row
        fill = _CONFIDENCE_FILLS.get(confidence_raw)
        if fill:
            for cell in ws_a[row_idx]:
                cell.fill = fill

    _auto_size_columns(ws_a, ASSUMPTIONS_HEADERS)

    # ------------------------------------------------------------------
    # Sheet 2 — Summary
    # ------------------------------------------------------------------
    _write_header(ws_s, SUMMARY_HEADERS, _HEADER_FILL, _HEADER_FONT)

    total = len(assumptions)
    confidence_counts: Counter[str] = Counter(
        str(a.get("confidence") or "").lower() for a in assumptions
    )
    high_count = confidence_counts.get("high", 0)
    medium_count = confidence_counts.get("medium", 0)
    low_count = confidence_counts.get("low", 0)

    vm_names = [str(a.get("vm_name") or "") for a in assumptions]
    unique_vms = sorted(set(v for v in vm_names if v))
    total_vms = len(unique_vms)
    avg_per_vm = round(total / total_vms, 1) if total_vms > 0 else 0.0

    # Top 5 VMs by assumption count
    vm_counter: Counter[str] = Counter(vm_names)
    top_5 = ", ".join(vm for vm, _ in vm_counter.most_common(5) if vm)

    summary_rows = [
        ("Total Assumptions", total),
        ("High Confidence", high_count),
        ("Medium Confidence", medium_count),
        ("Low Confidence", low_count),
        ("Total VMs/Servers", total_vms),
        ("Average Assumptions per VM", avg_per_vm),
        ("VMs with Most Assumptions", top_5),
    ]

    for metric, value in summary_rows:
        ws_s.append([metric, value])
        # Bold the metric label
        ws_s.cell(ws_s.max_row, 1).font = _BOLD_FONT

    _auto_size_columns(ws_s, SUMMARY_HEADERS)

    # ------------------------------------------------------------------
    # Sheet 3 — Excluded Servers (only when caller provides the list)
    # ------------------------------------------------------------------
    if excluded_servers:
        ws_e = wb.create_sheet("Excluded Servers")
        _write_header(ws_e, EXCLUDED_SERVERS_HEADERS, _HEADER_FILL, _HEADER_FONT)
        for rec in excluded_servers:
            timestamp = rec.get("excluded_at")
            if timestamp is not None and hasattr(timestamp, "isoformat"):
                timestamp = timestamp.isoformat()
            row_data = [
                rec.get("vm_name"),
                rec.get("os_config"),
                rec.get("server_type"),
                rec.get("exclusion_reason") or "(no reason provided)",
                timestamp,
            ]
            ws_e.append(row_data)
            for cell in ws_e[ws_e.max_row]:
                cell.fill = _ORANGE_FILL
        _auto_size_columns(ws_e, EXCLUDED_SERVERS_HEADERS)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
